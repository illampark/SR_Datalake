"""
Pipeline Engine — MQTT 기반 실시간 데이터 파이프라인 엔진

흐름:
  1. 파이프라인 시작 → PipelineBinding에 따라 MQTT 토픽 구독
  2. 메시지 수신 → PipelineStep 순서대로 모듈 체인 실행
  3. 처리 결과 → output 토픽에 발행 + 메타데이터 갱신
"""

import json
import logging
import time
import copy
from datetime import datetime, timedelta

from backend.services import mqtt_manager
from backend.services.pipeline_modules import MODULE_REGISTRY, SINK_REGISTRY, process_message, flush_all_sink_buffers
from backend.services import metadata_tracker

# 파일 소스 step 타입 — start_pipeline 시 MQTT 구독을 건너뛰고 트리거 기반으로 동작
_FILE_SOURCE_TYPES = {"import_source", "internal_file_source"}

# 파일 확장자 → 연결자 import_type 매핑. run_file_source 에서 연결자 설정과
# 실제 파일 확장자가 다를 때 fail-fast 하기 위한 용도.
_EXT_TO_IMPORT_TYPE = {
    "xlsx": "xlsx", "xls": "xlsx",
    "json": "json",
    "csv": "csv", "tsv": "csv", "txt": "csv",
}

logger = logging.getLogger(__name__)

# 실행 중인 파이프라인 상태
_running_pipelines = {}  # pipeline_id → {"steps": [...], "bindings": [...], "stats": {...}}

# ── 커넥터명 캐시 ──
_connector_name_cache = {}


def _resolve_connector_name(connector_type, connector_id):
    """커넥터 타입+ID로 커넥터 이름 조회 (캐시)"""
    key = (connector_type, connector_id)
    if key in _connector_name_cache:
        return _connector_name_cache[key]

    from backend.database import SessionLocal
    from backend.models.collector import (
        OpcuaConnector, ModbusConnector,
        MqttConnector, ApiConnector, FileCollector, DbConnector,
    )
    MODEL_MAP = {
        "opcua": OpcuaConnector,
        "modbus": ModbusConnector, "mqtt": MqttConnector,
        "api": ApiConnector, "file": FileCollector, "db": DbConnector,
    }
    model_cls = MODEL_MAP.get(connector_type)
    name = ""
    if model_cls:
        db = SessionLocal()
        try:
            row = db.query(model_cls.name).filter(model_cls.id == connector_id).first()
            name = row[0] if row else ""
        finally:
            db.close()
    _connector_name_cache[key] = name
    return name


def start_pipeline(pipeline_id):
    """파이프라인 시작 — DB에서 설정 로드 후 MQTT 구독"""
    from backend.database import SessionLocal
    from backend.models.pipeline import Pipeline, PipelineStep, PipelineBinding

    if pipeline_id in _running_pipelines:
        logger.info("파이프라인 %s 이미 실행 중", pipeline_id)
        return True

    db = SessionLocal()
    try:
        p = db.query(Pipeline).get(pipeline_id)
        if not p:
            logger.error("파이프라인 %s 없음", pipeline_id)
            return False

        steps = db.query(PipelineStep).filter_by(
            pipeline_id=pipeline_id
        ).order_by(PipelineStep.step_order).all()

        bindings = db.query(PipelineBinding).filter_by(
            pipeline_id=pipeline_id, enabled=True
        ).all()

        # 파일 소스 step 이 있으면 binding 없어도 실행 가능
        _has_file_source_step = any(
            st.enabled and st.module_type in _FILE_SOURCE_TYPES for st in steps
        )
        if not bindings and not _has_file_source_step:
            logger.warning("바인딩 없음", extra={"pipeline_id": pipeline_id})
            return False

        # MQTT 연결 확인
        if not mqtt_manager.is_connected():
            mqtt_manager.connect()
            if not mqtt_manager.is_connected():
                logger.error("MQTT 브로커 연결 불가")
                return False

        # 스텝 정보 캐시 (소스/처리/싱크 분리). step_id 를 함께 보관해서 step 단위 통계 갱신.
        source_steps = []
        step_configs = []
        sink_configs = []
        step_stats = {}
        for st in steps:
            if not st.enabled:
                continue
            cfg = dict(st.config or {})
            cfg["_pipeline_id"] = pipeline_id
            cfg["_step_id"] = st.id
            cfg["_step_type"] = st.module_type
            entry = {"step_id": st.id, "module_type": st.module_type, "config": cfg}
            step_stats[st.id] = {"processed": 0, "errors": 0, "dropped": 0, "last_at": None}
            if st.module_type in _FILE_SOURCE_TYPES:
                source_steps.append(entry)
            elif st.module_type in SINK_REGISTRY:
                sink_configs.append(entry)
            else:
                step_configs.append(entry)

        is_file_source = bool(source_steps)
        source_mode = "file" if is_file_source else "mqtt"

        _running_pipelines[pipeline_id] = {
            "name": p.name,
            "sources": source_steps,
            "steps": step_configs,
            "sinks": sink_configs,
            "source_mode": source_mode,
            "bindings": [{"type": b.connector_type, "id": b.connector_id, "filter": b.tag_filter} for b in bindings],
            "stats": {
                "processed": 0,
                "errors": 0,
                "dropped": 0,
                "started_at": datetime.utcnow().isoformat(),
            },
            "step_stats": step_stats,
            # MQTT-source 주기적 DB commit 타임스탬프 (file-source 는 별도 로직)
            "_last_commit_ts": time.time(),
        }

        # MQTT-source 시작 시 step 카운터 DB 리셋 (file-source 는 run_file_source 내부에서 처리)
        if not is_file_source:
            try:
                _reset_step_progress(pipeline_id)
            except Exception as e:
                logger.warning("step 카운터 리셋 실패 — %s", e,
                               extra={"pipeline_id": pipeline_id,
                                      "exc_class": type(e).__name__})

        if is_file_source:
            logger.info(
                "파이프라인 %s (%s): 파일 소스 모드 — MQTT 구독 안함, run-file-source 트리거 대기 (%d processing steps, %d sinks)",
                pipeline_id, p.name, len(step_configs), len(sink_configs),
            )
        else:
            # 바인딩별 MQTT 토픽 구독
            for b in bindings:
                topic = f"sdl/raw/{b.connector_type}/{b.connector_id}/#"
                if b.tag_filter and b.tag_filter != "*":
                    topic = f"sdl/raw/{b.connector_type}/{b.connector_id}/{b.tag_filter}"

                def make_handler(pid):
                    def handler(t, payload):
                        _handle_message(pid, t, payload)
                    return handler

                mqtt_manager.subscribe(topic, make_handler(pipeline_id))
                logger.info("파이프라인 %s: 구독 %s", pipeline_id, topic)

            logger.info("파이프라인 %s (%s) 시작 — %d 스텝, %d 바인딩",
                         pipeline_id, p.name, len(step_configs), len(bindings))
        return True
    except Exception as e:
        logger.error("시작 실패: %s", e,
                     extra={"pipeline_id": pipeline_id,
                            "exc_class": type(e).__name__})
        return False
    finally:
        db.close()


def stop_pipeline(pipeline_id):
    """파이프라인 정지 — MQTT 구독 해제"""
    info = _running_pipelines.pop(pipeline_id, None)
    if not info:
        return

    # 싱크 버퍼 플러시 (배치 모드 잔여 데이터 기록)
    try:
        flush_all_sink_buffers()
    except Exception as e:
        logger.warning("싱크 버퍼 플러시 실패 — %s", e,
                       extra={"pipeline_id": pipeline_id,
                              "exc_class": type(e).__name__})

    # 바인딩별 토픽 구독 해제
    for b in info.get("bindings", []):
        topic = f"sdl/raw/{b['type']}/{b['id']}/#"
        mqtt_manager.unsubscribe(topic)

    logger.info("파이프라인 %s 정지", pipeline_id)

    # DB 상태 갱신 + 마지막 step 카운터 flush
    _update_pipeline_db(pipeline_id, "stopped", info["stats"])
    try:
        _commit_step_progress(info.get("step_stats", {}))
    except Exception as e:
        logger.warning("stop 시 step 통계 flush 실패 — %s", e,
                       extra={"pipeline_id": pipeline_id,
                              "exc_class": type(e).__name__})


def get_pipeline_status(pipeline_id):
    """파이프라인 런타임 상태"""
    info = _running_pipelines.get(pipeline_id)
    if not info:
        return {"running": False}
    return {
        "running": True,
        "name": info["name"],
        "stats": info["stats"],
        "stepCount": len(info["steps"]),
        "bindingCount": len(info["bindings"]),
    }


def get_all_status():
    """모든 실행 중 파이프라인 상태"""
    result = {}
    for pid, info in _running_pipelines.items():
        result[pid] = {
            "name": info["name"],
            "stats": info["stats"],
        }
    return result


# ══════════════════════════════════════════════
# Internal: Message Processing
# ══════════════════════════════════════════════

def _handle_message(pipeline_id, topic, payload):
    """MQTT 메시지 수신 → 모듈 체인 실행"""
    info = _running_pipelines.get(pipeline_id)
    if not info:
        return

    start_time = time.time()
    step_stats = info.get("step_stats", {})
    now_dt = datetime.utcnow()

    try:
        # 페이로드 파싱
        if isinstance(payload, str):
            message = json.loads(payload)
        elif isinstance(payload, bytes):
            message = json.loads(payload.decode("utf-8"))
        else:
            message = payload

        # 원본 보존
        original = copy.deepcopy(message)

        # 소스 step (opcua_source / modbus_source / api_source / ...) — MQTT 메시지 도달 시점에
        # "수신 1건" 으로 카운트. 실제 소스 모듈은 stub 이라 process_message 안에서 카운트 못 함.
        for src in info.get("sources", []):
            sid = src.get("step_id")
            if sid and sid in step_stats:
                step_stats[sid]["processed"] += 1
                step_stats[sid]["last_at"] = now_dt

        # 모듈 체인 실행
        steps_applied = []
        for step in info["steps"]:
            module_type = step["module_type"]
            config = step["config"]
            sid = step.get("step_id")

            try:
                message = process_message(message, module_type, config)
            except Exception:
                if sid and sid in step_stats:
                    step_stats[sid]["errors"] += 1
                    step_stats[sid]["last_at"] = now_dt
                raise
            if message is None:
                # 메시지 드롭 — 현재 step 의 dropped 만 증가
                if sid and sid in step_stats:
                    step_stats[sid]["dropped"] += 1
                    step_stats[sid]["last_at"] = now_dt
                info["stats"]["dropped"] += 1
                return

            if sid and sid in step_stats:
                step_stats[sid]["processed"] += 1
                step_stats[sid]["last_at"] = now_dt
            steps_applied.append(module_type)

        # 처리 완료 → output 토픽 발행
        source = message.get("source", {})
        tag_name = source.get("tagName", "unknown")
        output_topic = mqtt_manager.build_processed_topic(pipeline_id, tag_name)
        mqtt_manager.publish(output_topic, message)

        # 싱크 실행 (처리 완료된 메시지를 각 싱크로 전달)
        for sink_step in info.get("sinks", []):
            sid = sink_step.get("step_id")
            try:
                sink_func = SINK_REGISTRY.get(sink_step["module_type"])
                if sink_func:
                    sink_func(copy.deepcopy(message), sink_step["config"])
                    if sid and sid in step_stats:
                        step_stats[sid]["processed"] += 1
                        step_stats[sid]["last_at"] = now_dt
            except Exception as se:
                if sid and sid in step_stats:
                    step_stats[sid]["errors"] += 1
                    step_stats[sid]["last_at"] = now_dt
                logger.warning("싱크 %s 오류 — %s",
                               sink_step["module_type"], se,
                               extra={"pipeline_id": pipeline_id,
                                      "step_id": sid,
                                      "step_type": sink_step["module_type"],
                                      "exc_class": type(se).__name__})

        # 통계 갱신
        info["stats"]["processed"] += 1
        elapsed_ms = (time.time() - start_time) * 1000

        # MQTT-source 주기적 DB commit (5초마다 또는 1000건마다)
        # file-source 는 run_file_source 내부에서 별도로 commit 함
        if info.get("source_mode") == "mqtt":
            now_ts = time.time()
            if (now_ts - info.get("_last_commit_ts", 0)) >= 5.0 \
                    or info["stats"]["processed"] % 1000 == 0:
                info["_last_commit_ts"] = now_ts
                # stop 과의 race 방지: _running_pipelines 에서 빠졌다면 commit X
                # (stop 이 status='stopped' 로 write 한 직후 race 로 'running' 으로
                # 되돌리는 사고 방지)
                if pipeline_id in _running_pipelines:
                    try:
                        _update_pipeline_db(pipeline_id, "running", info["stats"])
                        _commit_step_progress(step_stats)
                    except Exception as ce:
                        logger.warning("주기적 통계 commit 실패 — %s", ce,
                                       extra={"pipeline_id": pipeline_id,
                                              "exc_class": type(ce).__name__})

        # 메타데이터 갱신 (비동기적)
        try:
            metadata_tracker.upsert_tag_metadata(
                connector_type=source.get("connectorType", ""),
                connector_id=source.get("connectorId", 0),
                connector_name=_resolve_connector_name(
                    source.get("connectorType", ""),
                    source.get("connectorId", 0),
                ),
                tag_name=tag_name,
                value=message.get("value"),
                data_type=message.get("dataType", "float"),
                unit=message.get("unit", ""),
            )

            # 계보 기록 (100건 배치 단위)
            if info["stats"]["processed"] % 100 == 0:
                metadata_tracker.record_lineage(
                    source_connector_type=source.get("connectorType", ""),
                    source_connector_id=source.get("connectorId", 0),
                    source_tag=tag_name,
                    pipeline_id=pipeline_id,
                    pipeline_name=info["name"],
                    steps_applied=steps_applied,
                    destination_type="mqtt",
                    destination_target=output_topic,
                    record_count=100,
                    processing_ms=elapsed_ms,
                )
        except Exception as e:
            logger.warning("메타데이터 갱신 실패: %s", e)

    except json.JSONDecodeError as je:
        info["stats"]["errors"] += 1
        logger.warning("JSON 파싱 오류 [%s]", topic,
                       extra={"pipeline_id": pipeline_id,
                              "exc_class": "JSONDecodeError"})
    except Exception as e:
        info["stats"]["errors"] += 1
        logger.error("처리 오류 [%s] — %s", topic, e,
                     extra={"pipeline_id": pipeline_id,
                            "exc_class": type(e).__name__})

        # DB에 에러 기록
        _update_pipeline_error(pipeline_id, str(e))


def get_file_source_lock_state(pipeline_id):
    """파이프라인의 file-source 락 상태를 동기 조회.

    Returns: (is_locked: bool, run_id: str|None, run_at: datetime|None).
    24h 이상 지나면 stale 으로 보고 is_locked=False (다음 acquire 가 알아서 빼앗음).
    """
    from backend.database import SessionLocal
    from sqlalchemy import text as _t
    ss = SessionLocal()
    try:
        row = ss.execute(_t(
            "SELECT current_run_id, current_run_at FROM pipeline WHERE id=:p"
        ), {"p": pipeline_id}).fetchone()
        if not row:
            return (False, None, None)
        rid, rat = row[0], row[1]
        if not rid:
            return (False, None, None)
        if rat is None:
            return (False, rid, None)
        if (datetime.utcnow() - rat) > timedelta(hours=24):
            return (False, rid, rat)
        return (True, rid, rat)
    finally:
        ss.close()


def clear_stale_file_source_locks():
    """컨테이너 시작 시 호출 — 모든 파이프라인의 current_run_id 를 NULL 로 리셋.

    컨테이너가 재시작되면 in-flight file-source 스레드는 사라진 상태이므로 어떤 lock
    도 유효할 수 없다. 이대로 두면 사용자가 /run-file-source 를 호출해도 stale lock 에
    막혀 24h 동안 처리가 안 됨 (`catalog_export_worker.resume_pending_on_startup` 과
    동일한 클래스의 startup-cleanup 문제).
    """
    from backend.database import SessionLocal
    from sqlalchemy import text as _t
    ss = SessionLocal()
    try:
        result = ss.execute(_t("""
            UPDATE pipeline
            SET current_run_id = '', current_run_at = NULL
            WHERE current_run_id IS NOT NULL AND current_run_id <> ''
        """))
        ss.commit()
        cleared = result.rowcount or 0
        if cleared:
            logger.info("startup: cleared %d stale file-source lock(s)", cleared)
        return cleared
    except Exception as e:
        logger.warning("startup: stale lock cleanup failed — %s", e)
        ss.rollback()
        return 0
    finally:
        ss.close()


def _reset_step_progress(pipeline_id):
    """파이프라인 시작 시 pipeline_step 카운터 0으로 리셋."""
    from backend.database import SessionLocal
    from sqlalchemy import text as _t
    ss = SessionLocal()
    try:
        ss.execute(_t("""
            UPDATE pipeline_step
            SET processed_count = 0, error_count = 0, dropped_count = 0,
                last_processed_at = NULL
            WHERE pipeline_id = :pid
        """), {"pid": pipeline_id})
        ss.commit()
    except Exception:
        ss.rollback()
    finally:
        ss.close()


def _commit_step_progress(step_stats):
    """step_stats: {step_id: {processed, errors, dropped, last_at}}.
    유효한 step_id 항목만 골라 UPDATE."""
    if not step_stats:
        return
    from backend.database import SessionLocal
    from sqlalchemy import text as _t
    ss = SessionLocal()
    try:
        for sid, s in step_stats.items():
            if not sid:
                continue
            ss.execute(_t("""
                UPDATE pipeline_step
                SET processed_count = :p, error_count = :e, dropped_count = :d,
                    last_processed_at = :ts
                WHERE id = :sid
            """), {
                "p": s.get("processed", 0),
                "e": s.get("errors", 0),
                "d": s.get("dropped", 0),
                "ts": s.get("last_at"),
                "sid": sid,
            })
        ss.commit()
    except Exception:
        ss.rollback()
    finally:
        ss.close()


def _update_pipeline_db(pipeline_id, status, stats):
    """파이프라인 DB 상태 갱신"""
    from backend.database import SessionLocal
    from backend.models.pipeline import Pipeline
    db = SessionLocal()
    try:
        p = db.query(Pipeline).get(pipeline_id)
        if p:
            p.status = status
            p.processed_count = stats.get("processed", 0)
            p.error_count = stats.get("errors", 0)
            p.last_processed_at = datetime.utcnow()
            db.commit()
    except Exception as e:
        db.rollback()
        logger.error("DB 갱신 실패: %s", e,
                     extra={"pipeline_id": pipeline_id,
                            "exc_class": type(e).__name__})
    finally:
        db.close()


def _update_pipeline_error(pipeline_id, error_msg):
    """파이프라인 에러 기록"""
    from backend.database import SessionLocal
    from backend.models.pipeline import Pipeline
    db = SessionLocal()
    try:
        p = db.query(Pipeline).get(pipeline_id)
        if p:
            p.last_error = error_msg
            p.error_count = (p.error_count or 0) + 1
            db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()





def _resume_load(pipeline_id):
    """파이프라인의 resume_state 불러오기. 없으면 빈 dict."""
    from backend.database import SessionLocal
    from sqlalchemy import text as _t
    ss = SessionLocal()
    try:
        row = ss.execute(_t("SELECT resume_state FROM pipeline WHERE id=:p"),
                         {"p": pipeline_id}).fetchone()
        return (row[0] if row and row[0] else {}) or {}
    except Exception as e:
        logger.warning("resume_state load failed pid=%s: %s", pipeline_id, e)
        return {}
    finally:
        ss.close()


def _resume_save(pipeline_id, state):
    """resume_state 저장 — 진행 중간 또는 정지 직전 호출."""
    from backend.database import SessionLocal
    from sqlalchemy import text as _t
    ss = SessionLocal()
    try:
        ss.execute(_t("UPDATE pipeline SET resume_state = cast(:s as json) WHERE id=:p"),
                   {"s": json.dumps(state, ensure_ascii=False), "p": pipeline_id})
        ss.commit()
    except Exception as e:
        ss.rollback()
        logger.warning("resume_state save failed pid=%s: %s", pipeline_id, e)
    finally:
        ss.close()


def _resume_clear(pipeline_id):
    """파이프라인이 정상 완료됐을 때 resume_state 비우기."""
    from backend.database import SessionLocal
    from sqlalchemy import text as _t
    ss = SessionLocal()
    try:
        ss.execute(_t("UPDATE pipeline SET resume_state = NULL WHERE id=:p"),
                   {"p": pipeline_id})
        ss.commit()
    except Exception:
        ss.rollback()
    finally:
        ss.close()


def _is_stop_requested(pipeline_id):
    """현재 pipeline.status 가 'stopped' 면 외부에서 정지 요청한 것."""
    from backend.database import SessionLocal
    from sqlalchemy import text as _t
    ss = SessionLocal()
    try:
        st = ss.execute(_t("SELECT status FROM pipeline WHERE id=:p"),
                        {"p": pipeline_id}).scalar()
        return (str(st or "").lower() == "stopped")
    except Exception:
        return False
    finally:
        ss.close()


def run_file_source(pipeline_id):
    """파일 소스 파이프라인 즉시 실행 — DB lock + 스트리밍 + 진행률 commit + 재개(resume).

    DB 의 current_run_id 를 atomic UPDATE 로 set 하여 동일 파이프라인의
    동시 실행을 차단한다 (24h TTL 로 crash 후 stale lock 자동 해제).
    CSV 는 chunk 단위 line iterator + csv.reader 로, JSON 은 ijson 으로
    스트리밍 처리해 큰 파일도 메모리에 통째 적재하지 않는다.
    """
    import csv as _csv
    import time as _time
    import uuid as _uuid
    from datetime import timedelta as _timedelta
    from sqlalchemy import text as _sql_text
    from backend.database import SessionLocal
    from backend.models.pipeline import Pipeline, PipelineStep
    from backend.models.collector import ImportCollector
    from backend.services.minio_client import get_minio_client
    from backend.services.import_parser import _parse_json
    from backend.services.pipeline_modules import flush_all_sink_buffers

    PROGRESS_RECORDS = 5000
    PROGRESS_SECONDS = 5
    LOCK_TTL_HOURS = 24

    def _commit_progress(processed, errors, dropped):
        ss = SessionLocal()
        try:
            pp = ss.query(Pipeline).get(pipeline_id)
            if pp:
                pp.processed_count = processed
                pp.error_count = errors
                pp.last_processed_at = datetime.utcnow()
                ss.commit()
        except Exception:
            ss.rollback()
        finally:
            ss.close()

    def _reset_step_progress():
        ss = SessionLocal()
        try:
            ss.execute(_sql_text("""
                UPDATE pipeline_step
                SET processed_count = 0, error_count = 0, dropped_count = 0,
                    last_processed_at = NULL
                WHERE pipeline_id = :pid
            """), {"pid": pipeline_id})
            ss.commit()
        except Exception:
            ss.rollback()
        finally:
            ss.close()

    def _commit_step_progress(step_stats):
        if not step_stats:
            return
        ss = SessionLocal()
        try:
            for sid, s in step_stats.items():
                ss.execute(_sql_text("""
                    UPDATE pipeline_step
                    SET processed_count = :p, error_count = :e, dropped_count = :d,
                        last_processed_at = :ts
                    WHERE id = :sid
                """), {
                    "p": s["processed"], "e": s["errors"], "d": s["dropped"],
                    "ts": s["last_at"], "sid": sid,
                })
            ss.commit()
        except Exception as e:
            ss.rollback()
            logger.warning("file-source _commit_step_progress 실패: %s", e,
                           extra={"pipeline_id": pipeline_id,
                                  "exc_class": type(e).__name__})
        finally:
            ss.close()

    def _try_acquire_lock(run_id):
        ss = SessionLocal()
        try:
            now = datetime.utcnow()
            cutoff = now - _timedelta(hours=LOCK_TTL_HOURS)
            res = ss.execute(_sql_text("""
                UPDATE pipeline
                SET current_run_id = :rid, current_run_at = :now
                WHERE id = :pid
                  AND (current_run_id IS NULL OR current_run_id = ''
                       OR current_run_at IS NULL OR current_run_at < :cutoff)
                RETURNING current_run_id
            """), {"rid": run_id, "now": now, "cutoff": cutoff, "pid": pipeline_id})
            row = res.fetchone()
            ss.commit()
            return bool(row) and row[0] == run_id
        except Exception:
            ss.rollback()
            return False
        finally:
            ss.close()

    def _release_lock(run_id):
        ss = SessionLocal()
        try:
            ss.execute(_sql_text("""
                UPDATE pipeline
                SET current_run_id = '', current_run_at = NULL
                WHERE id = :pid AND current_run_id = :rid
            """), {"pid": pipeline_id, "rid": run_id})
            ss.commit()
        except Exception:
            ss.rollback()
        finally:
            ss.close()

    def _stream_csv_records(resp, encoding, delimiter, skip_header,
                             stream_base_byte=0, fixed_header=None):
        """CSV 스트리밍.

        stream_base_byte: 이 스트림이 파일의 어느 byte 위치에서 시작하는지.
                          MinIO Range 로 부분 GET 했을 때 caller 가 base 를 전달한다.
        fixed_header:    Range 부분 GET 의 경우 헤더가 스트림에 없으므로 외부에서
                          미리 파싱한 헤더 (list[str]) 를 주입한다.

        yields (record_dict, abs_byte_after_line, header) 튜플.
          - abs_byte_after_line = stream_base_byte + 이 스트림에서 소비한 bytes
            (직전 line 의 \n 직후, 다음 line 시작 위치 = 안전한 resume offset)
          - header 는 이 스트림에서 처음 추출했을 때만 의미 있음. caller 가
            resume_state 에 저장해두면 다음 run 에서 fixed_header 로 재주입.
        """
        state = {"bytes": 0}  # 이 stream 에서 누적 소비 bytes (\n 포함)
        first_chunk = [True]

        def _line_iter():
            buf = b""
            for chunk in resp.stream(1 * 1024 * 1024):
                if not isinstance(chunk, (bytes, bytearray)):
                    continue
                if first_chunk[0]:
                    first_chunk[0] = False
                    # offset=0 인 경우에만 BOM 가능성 있음 (Range 요청은 BOM 미포함)
                    if stream_base_byte == 0 and chunk.startswith(b"\xef\xbb\xbf"):
                        state["bytes"] += 3
                        chunk = chunk[3:]
                buf += chunk
                while True:
                    nl = buf.find(b"\n")
                    if nl < 0:
                        break
                    line_b = buf[:nl]
                    buf = buf[nl + 1:]
                    state["bytes"] += nl + 1  # line + \n
                    if line_b.endswith(b"\r"):
                        line_b = line_b[:-1]
                    yield line_b.decode(encoding, errors="replace")
            if buf:
                state["bytes"] += len(buf)
                if buf.endswith(b"\r"):
                    buf = buf[:-1]
                yield buf.decode(encoding, errors="replace")

        reader = _csv.reader(_line_iter(), delimiter=delimiter)
        header = list(fixed_header) if fixed_header else None
        for row in reader:
            if header is None:
                if skip_header:
                    header = [c.strip() for c in row]
                    continue
                else:
                    header = [f"col{i}" for i in range(len(row))]
            if len(row) == len(header):
                rec = dict(zip(header, row))
            else:
                rec = {h: (row[i] if i < len(row) else "") for i, h in enumerate(header)}
            yield rec, stream_base_byte + state["bytes"], header

    def _stream_json_records(resp):
        """ijson 으로 JSON 스트리밍. 루트가 array (item) 또는 {data|records|items: [...]} 지원."""
        try:
            import ijson
        except ImportError:
            logger.warning("ijson 미설치 — JSON 전체 메모리 적재 fallback")
            content = resp.read()
            for r in _parse_json(content, "utf-8"):
                yield r
            return

        # ijson 은 'item' (루트 배열) / 'data.item' 등 prefix 지원.
        # 일반적으로 사용자 JSON 은 [...] 또는 {"data": [...]}, {"records": [...]}.
        # 우선 'item' 을 시도하고, 안되면 fallback 으로 일반 parse.
        try:
            for rec in ijson.items(resp, "item"):
                if isinstance(rec, dict):
                    yield rec
                else:
                    yield {"value": rec}
        except Exception as e:
            logger.warning("ijson 스트리밍 실패 (%s) — 메모리 적재 fallback", e)
            try:
                resp.close(); resp.release_conn()
            except Exception:
                pass
            # ijson 이 stream 을 소비했을 수 있으므로 재시도는 별도 get_object 가 필요.
            # 호출자가 처리해야 함. 빈 yield.
            return

    def _stream_xlsx_records(resp, sheet_name, header_row):
        """xlsx 스트리밍. ZIP+시트 XML 구조라 진정한 streaming 은 불가능 →
        BytesIO 적재 후 openpyxl read_only 모드로 행 단위 yield.
        업무용 수십 MB 범위는 안전. GB 급은 권장하지 않음.
        """
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl 미설치 — xlsx 처리 불가")
            try:
                resp.close(); resp.release_conn()
            except Exception:
                pass
            return

        from io import BytesIO
        try:
            buf = BytesIO(resp.read())
        finally:
            try:
                resp.close(); resp.release_conn()
            except Exception:
                pass

        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb[wb.sheetnames[0]]

            hr = max(1, int(header_row or 1))
            headers = None
            for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if ri < hr:
                    continue
                if ri == hr:
                    headers = [(str(c).strip() if c is not None else f"col_{i}") for i, c in enumerate(row)]
                    continue
                if headers is None:
                    continue
                if all(c is None for c in row):
                    continue
                rec = {}
                for i, h in enumerate(headers):
                    v = row[i] if i < len(row) else None
                    # 파이프라인 sink 호환 위해 datetime → ISO 문자열로 변환 (None/숫자/문자열은 그대로)
                    if hasattr(v, "isoformat"):
                        v = v.isoformat()
                    rec[h] = v
                yield rec
        finally:
            wb.close()

    db = SessionLocal()
    try:
        p = db.query(Pipeline).get(pipeline_id)
        if not p:
            return {"ok": False, "error": f"Pipeline {pipeline_id} 를 찾을 수 없습니다."}
        if p.status != "running":
            return {"ok": False, "error": "파이프라인이 실행 중이 아닙니다. 먼저 [시작]을 누르세요."}

        # 동시 실행 방지 — DB 기반 lock
        run_id = str(_uuid.uuid4())
        if not _try_acquire_lock(run_id):
            return {
                "ok": False,
                "error": "이미 실행 중인 file-source 작업이 있습니다. 완료를 기다린 뒤 재시도하세요.",
            }

        try:
            steps = db.query(PipelineStep).filter_by(
                pipeline_id=pipeline_id
            ).order_by(PipelineStep.step_order).all()

            source_steps = []
            process_steps = []
            sink_steps = []
            step_stats = {}  # step_id → {"processed", "errors", "dropped", "last_at"}
            for st in steps:
                if not st.enabled:
                    continue
                entry = {"step_id": st.id, "module_type": st.module_type, "config": dict(st.config or {})}
                step_stats[st.id] = {"processed": 0, "errors": 0, "dropped": 0, "last_at": None}
                if st.module_type in _FILE_SOURCE_TYPES:
                    source_steps.append(entry)
                elif st.module_type in SINK_REGISTRY:
                    sink_steps.append(entry)
                else:
                    process_steps.append(entry)
            source_step_id = source_steps[0].get("step_id") if source_steps else None

            if not source_steps:
                return {"ok": False, "error": "import_source 스텝이 없습니다."}

            src_cfg = source_steps[0]["config"]
            connector_type = src_cfg.get("connectorType", "import")
            connector_id = src_cfg.get("connectorId")
            if not connector_id:
                return {"ok": False, "error": "source 설정에 connectorId 가 없습니다."}

            c = db.query(ImportCollector).get(connector_id)
            if not c:
                return {"ok": False, "error": f"Import collector {connector_id} 를 찾을 수 없습니다."}

            client = get_minio_client(db)
            bucket = c.target_bucket or "sdl-files"
            prefix = f"import/{connector_id}/"

            objects = list(client.list_objects(bucket, prefix=prefix, recursive=True))
            if not objects:
                return {"ok": False, "error": f"s3://{bucket}/{prefix} 에 파일이 없습니다."}

            encoding = c.encoding or "utf-8"
            delimiter = c.delimiter or ","
            skip_header = c.skip_header
            import_type = (c.import_type or "csv").lower()

            # ── 재개 상태 복원 ──
            _resume_state = _resume_load(pipeline_id)
            _processed_objects = set(_resume_state.get("processed_objects") or [])
            _resume_object = _resume_state.get("current_object") or None
            _resume_record_idx = int(_resume_state.get("current_record_idx") or 0)
            _resume_byte_offset = int(_resume_state.get("current_byte_offset") or 0)
            _resume_header = _resume_state.get("current_header") or None
            _resume_processed = int(_resume_state.get("processed_count_at_save") or 0)
            _resume_errors = int(_resume_state.get("error_count_at_save") or 0)
            _resume_dropped = int(_resume_state.get("dropped_count_at_save") or 0)

            if _processed_objects or _resume_object:
                logger.info(
                    "file-source pipeline=%s 재개 — completed_files=%d, resume_in='%s' "
                    "byte_offset=%d record_idx=%d",
                    pipeline_id, len(_processed_objects), _resume_object,
                    _resume_byte_offset, _resume_record_idx,
                )

            logger.info(
                "file-source pipeline=%s 시작 — files=%d (skip=%d), format=%s, sinks=%d, run_id=%s",
                pipeline_id, len(objects), len(_processed_objects), import_type, len(sink_steps), run_id,
            )

            # 카운터: 처음 시작이면 0, 재개면 직전 저장값에서 이어감
            total_processed = _resume_processed
            total_errors = _resume_errors
            total_dropped = _resume_dropped
            last_commit_at = _time.time()
            last_commit_count = total_processed

            # 재개 시작 시 누적 카운터 즉시 반영
            if _resume_processed or _resume_errors or _resume_dropped:
                _commit_progress(total_processed, total_errors, total_dropped)
            else:
                _commit_progress(0, 0, 0)
                _reset_step_progress()

            cancelled = False

            for fi, obj in enumerate(objects):
                # 이미 완료된 파일 skip
                if obj.object_name in _processed_objects:
                    logger.info("file-source pipeline=%s skip already-processed: %s",
                                pipeline_id, obj.object_name)
                    continue
                if cancelled:
                    break
                try:
                    _ext = obj.object_name.rsplit(".", 1)[-1].lower() if "." in obj.object_name else ""
                    _detected = _EXT_TO_IMPORT_TYPE.get(_ext)
                    if _detected and _detected != import_type:
                        raise ValueError(
                            f"연결자 형식이 '{import_type}'(으)로 설정돼 있으나 파일 확장자는 '.{_ext}' 입니다. "
                            f"수집 관리 → 가져오기 수집기에서 형식을 '{_detected}'(으)로 변경하세요."
                        )
                    # CSV 재개 시 byte offset 사용 — Range 로 부분 GET, 헤더는 saved 사용
                    is_resume_target = (obj.object_name == _resume_object)
                    use_byte_resume = (
                        import_type == "csv" and is_resume_target
                        and _resume_byte_offset > 0 and _resume_header
                    )
                    if use_byte_resume:
                        resp = client.get_object(
                            bucket, obj.object_name, offset=_resume_byte_offset,
                        )
                        logger.info(
                            "file-source pipeline=%s resume %s @ byte %d (skip parsing)",
                            pipeline_id, obj.object_name, _resume_byte_offset,
                        )
                    else:
                        resp = client.get_object(bucket, obj.object_name)
                    try:
                        # CSV 인 경우 byte-offset 모드 / 그 외 (json/xlsx) 는 record_idx 모드 유지
                        csv_byte_mode = (import_type == "csv")
                        if import_type == "json":
                            records_iter = _stream_json_records(resp)
                        elif import_type == "xlsx":
                            records_iter = _stream_xlsx_records(
                                resp, c.sheet_name, c.header_row,
                            )
                        else:
                            records_iter = _stream_csv_records(
                                resp, encoding, delimiter, skip_header,
                                stream_base_byte=(_resume_byte_offset if use_byte_resume else 0),
                                fixed_header=(_resume_header if use_byte_resume else None),
                            )

                        # 재개 대상 파일이면 첫 N record skip (byte resume 면 0)
                        if csv_byte_mode and use_byte_resume:
                            skip_records = 0
                            file_record_count = _resume_record_idx
                        elif is_resume_target:
                            skip_records = _resume_record_idx
                            if skip_records:
                                logger.info(
                                    "file-source pipeline=%s resume %s @ record %d (legacy skip)",
                                    pipeline_id, obj.object_name, skip_records,
                                )
                            file_record_count = 0
                        else:
                            skip_records = 0
                            file_record_count = 0
                        cur_byte_offset = _resume_byte_offset if use_byte_resume else 0
                        cur_header = _resume_header if use_byte_resume else None
                        skipped = 0
                        for _item in records_iter:
                            # CSV 는 (rec, abs_byte, header) tuple — 그 외는 dict 단일
                            if csv_byte_mode:
                                rec, cur_byte_offset, cur_header = _item
                            else:
                                rec = _item
                            if skipped < skip_records:
                                skipped += 1
                                continue
                            msg = {
                                "value": rec,
                                "source": {
                                    "connectorType": connector_type,
                                    "connectorId": connector_id,
                                    "tagName": "row",
                                },
                                "dataType": "object",
                                "timestamp": datetime.utcnow().isoformat(),
                            }

                            _now_ts = datetime.utcnow()
                            if source_step_id is not None and source_step_id in step_stats:
                                _s = step_stats[source_step_id]
                                _s["processed"] += 1
                                _s["last_at"] = _now_ts

                            drop = False
                            for step_entry in process_steps:
                                _sid = step_entry.get("step_id")
                                msg = process_message(
                                    msg, step_entry["module_type"], step_entry["config"],
                                )
                                if msg is None:
                                    if _sid in step_stats:
                                        step_stats[_sid]["dropped"] += 1
                                        step_stats[_sid]["last_at"] = _now_ts
                                    drop = True
                                    break
                                if _sid in step_stats:
                                    step_stats[_sid]["processed"] += 1
                                    step_stats[_sid]["last_at"] = _now_ts
                            if drop:
                                total_dropped += 1
                                continue

                            for sink_step in sink_steps:
                                _sid = sink_step.get("step_id")
                                try:
                                    sink_func = SINK_REGISTRY.get(sink_step["module_type"])
                                    if sink_func:
                                        cfg = dict(sink_step["config"])
                                        cfg["_pipeline_id"] = pipeline_id
                                        sink_func(copy.deepcopy(msg), cfg)
                                        if _sid in step_stats:
                                            step_stats[_sid]["processed"] += 1
                                            step_stats[_sid]["last_at"] = _now_ts
                                except Exception as se:
                                    total_errors += 1
                                    if _sid in step_stats:
                                        step_stats[_sid]["errors"] += 1
                                        step_stats[_sid]["last_at"] = _now_ts
                                    logger.warning(
                                        "file-source sink %s 실패: %s",
                                        sink_step["module_type"], se,
                                    )

                            total_processed += 1
                            file_record_count += 1

                            now = _time.time()
                            if (total_processed - last_commit_count) >= PROGRESS_RECORDS or (now - last_commit_at) >= PROGRESS_SECONDS:
                                _commit_progress(total_processed, total_errors, total_dropped)
                                _commit_step_progress(step_stats)
                                last_commit_count = total_processed
                                last_commit_at = now
                                # 재개 상태 저장 (CSV 는 byte offset 우선, 그 외는 record_idx)
                                _save_state = {
                                    "processed_objects": list(_processed_objects),
                                    "current_object": obj.object_name,
                                    "current_record_idx": skip_records + file_record_count,
                                    "processed_count_at_save": total_processed,
                                    "error_count_at_save": total_errors,
                                    "dropped_count_at_save": total_dropped,
                                }
                                if csv_byte_mode and cur_byte_offset > 0:
                                    _save_state["current_byte_offset"] = cur_byte_offset
                                    if cur_header:
                                        _save_state["current_header"] = cur_header
                                _resume_save(pipeline_id, _save_state)
                                # 외부에서 정지 요청? — 안전하게 break
                                if _is_stop_requested(pipeline_id):
                                    logger.info(
                                        "file-source pipeline=%s STOP 요청 감지 — 안전 정지 (file=%s record_idx=%d)",
                                        pipeline_id, obj.object_name,
                                        skip_records + file_record_count,
                                    )
                                    cancelled = True
                                    break
                    finally:
                        try:
                            resp.close(); resp.release_conn()
                        except Exception:
                            pass

                    if cancelled:
                        # 정지 요청 — break out without marking file as completed
                        break

                    logger.info(
                        "file-source pipeline=%s file=%s done — rows=%d (cumulative processed=%d)",
                        pipeline_id, obj.object_name, file_record_count, total_processed,
                    )
                    _commit_progress(total_processed, total_errors, total_dropped)
                    # 작은 파일이 빠르게 끝나는 경우 inner-loop 의 5000행/5초 임계치가
                    # 매 파일마다 리셋되어 step_stats 가 영영 commit 되지 않는 문제 방지 —
                    # 파일 단위 commit 도 함께 한다.
                    _commit_step_progress(step_stats)
                    last_commit_count = total_processed
                    last_commit_at = _time.time()
                    # 외부에서 정지 요청? — 작은 파일 처리 시 inner-loop 의 stop 체크가
                    # 절대 닿지 않으므로 (commit threshold reset 으로) 파일 경계에서도
                    # 명시적으로 확인. 다음 파일로 넘어가기 전에 안전 정지.
                    if _is_stop_requested(pipeline_id):
                        logger.info(
                            "file-source pipeline=%s STOP 요청 감지 (file boundary)",
                            pipeline_id,
                        )
                        cancelled = True
                    # 파일 정상 완료 → processed_objects 추가, 현재 파일 offset 리셋
                    _processed_objects.add(obj.object_name)
                    _resume_object = None
                    _resume_record_idx = 0
                    _resume_byte_offset = 0
                    _resume_header = None
                    _resume_save(pipeline_id, {
                        "processed_objects": list(_processed_objects),
                        "current_object": None,
                        "current_record_idx": 0,
                        "current_byte_offset": 0,
                        "current_header": None,
                        "processed_count_at_save": total_processed,
                        "error_count_at_save": total_errors,
                        "dropped_count_at_save": total_dropped,
                    })
                except Exception as e:
                    total_errors += 1
                    logger.warning("파일 %s 처리 실패: %s", obj.object_name, e)

            try:
                flush_all_sink_buffers()
            except Exception as e:
                logger.warning("flush_all_sink_buffers 실패: %s", e)

            _commit_progress(total_processed, total_errors, total_dropped)
            _commit_step_progress(step_stats)

            if cancelled:
                # 정지 요청 — resume_state 보존, 다음 실행 때 이어감
                logger.info(
                    "file-source pipeline=%s 정지됨 — files_done=%d processed=%d (다음 실행 시 재개)",
                    pipeline_id, len(_processed_objects), total_processed,
                )
                return {
                    "ok": True,
                    "cancelled": True,
                    "files_done": len(_processed_objects),
                    "processed": total_processed,
                    "dropped": total_dropped,
                    "errors": total_errors,
                }

            # 정상 완료 — resume_state 비우기
            _resume_clear(pipeline_id)

            logger.info(
                "file-source pipeline=%s 완료 — files=%d processed=%d dropped=%d errors=%d",
                pipeline_id, len(objects), total_processed, total_dropped, total_errors,
            )
            return {
                "ok": True,
                "files": len(objects),
                "processed": total_processed,
                "dropped": total_dropped,
                "errors": total_errors,
            }
        finally:
            _release_lock(run_id)
    finally:
        db.close()
