"""
Import Parser — CSV/JSON 파일 파싱 및 MQTT 발행을 통한 데이터 가져오기

Import Collector 아키텍처:
  파일 업로드 → 파싱 → MQTT 발행 (sdl/raw/import/{id}/{tag})
  → Pipeline 변환 (선택) → Storage Sink (TSDB/RDBMS/MinIO)
  → Catalog 자동 등록

직접 저장 모드 (publish_mqtt=False):
  파일 업로드 → 파싱 → 직접 INSERT (TSDB/RDBMS) 또는 MinIO 업로드
"""

import csv
import io
import json
import logging
import mimetypes
import os
import shutil
import threading
import zipfile
from datetime import datetime

logger = logging.getLogger(__name__)

_import_threads = {}  # collector_id → thread
_import_lock = threading.Lock()


def _apply_post_import_action(collector, full_path, rel_path):
    """import 성공 후 소스 파일 정리 — MinIO 정본화로 NFS·MinIO 중복 제거.

    collector.post_import_action:
      keep    — 소스 유지 (레거시 호환, 중복 발생)
      archive — {local_path}/{archive_subdir}/{rel_path} 로 이동 (무손실, 구조 보존)
      delete  — 소스 즉시 삭제

    full_path: 소스 파일 절대 경로
    rel_path : local_path 기준 상대 경로 (archive 시 구조 보존에 사용)

    실패는 import 전체를 중단시키지 않는다 — 파일은 이미 MinIO 에 안전히 적재됨.
    """
    action = (getattr(collector, "post_import_action", None) or "keep").lower()
    if action == "keep":
        return
    try:
        if action == "delete":
            os.remove(full_path)
            logger.info(f"Import #{collector.id} post-import delete: {rel_path}")
        elif action == "archive":
            subdir = collector.archive_subdir or ".imported"
            base = collector.local_path or os.path.dirname(full_path)
            dest = os.path.join(base, subdir, rel_path)
            os.makedirs(os.path.dirname(dest), exist_ok=True)
            shutil.move(full_path, dest)
            logger.info(f"Import #{collector.id} post-import archive: {rel_path} → {subdir}/")
    except Exception as e:
        logger.warning(
            f"Import #{collector.id} post-import '{action}' failed for {rel_path}: {e}"
        )


def _publish_file_meta(connector_type, connector_id, rel_path, bucket, object_name,
                       file_size, mime_type, tag_name="file"):
    """파이프라인 라우팅을 위한 file_meta envelope MQTT 발행.

    외부 파일 sink(external_file_sink) 가 기대하는 키 구조에 맞춤:
      value: {minio_bucket, minio_path, file_size, mime_type, original_path}
      dataType: "file_meta"

    파이프라인이 sdl/raw/{connector_type}/{connector_id}/# 를 구독 중이면
    이 메시지가 라우팅되어 외부 sink 가 MinIO 객체를 읽어 외부로 전송한다.
    """
    try:
        from backend.services import mqtt_manager
        topic = f"sdl/raw/{connector_type}/{connector_id}/{tag_name}"
        envelope = {
            "source": {
                "connectorType": connector_type,
                "connectorId": connector_id,
                "tagName": tag_name,
            },
            "value": {
                "minio_bucket": bucket,
                "minio_path": object_name,
                "file_size": file_size,
                "mime_type": mime_type,
                "original_path": rel_path,
            },
            "dataType": "file_meta",
            "timestamp": datetime.utcnow().isoformat() + "Z",
        }
        mqtt_manager.publish(topic, envelope, qos=1)
    except Exception as e:
        # 발행 실패가 import 전체를 중단시키지 않도록 — 파일 자체는 이미 MinIO 에 안전히 저장됨
        logger.warning(f"file_meta publish failed (cid={connector_id}, path={rel_path}): {e}")


def _build_rdbms_url(cfg):
    """RdbmsConfig로부터 SQLAlchemy 접속 URL 생성"""
    db_type = (cfg.db_type or "").lower()
    if "mysql" in db_type or "maria" in db_type:
        driver = "mysql+pymysql"
    else:
        driver = "postgresql+psycopg2"
    user = cfg.username or "sdl_user"
    pw = cfg.password or ""
    host = cfg.host or "localhost"
    port = cfg.port or 5432
    db = cfg.database_name or "sdl"
    return f"{driver}://{user}:{pw}@{host}:{port}/{db}"

# 인코딩 감지 순서 (한글 환경 우선)
_ENCODING_CANDIDATES = ["utf-8", "cp949", "euc-kr", "utf-8-sig", "latin-1"]


def _detect_and_decode(raw_bytes, hint_encoding="utf-8"):
    """바이트 데이터를 자동 인코딩 감지하여 문자열로 변환

    1) hint_encoding으로 시도
    2) 실패 시 후보 인코딩 순서대로 시도
    3) 최종 실패 시 latin-1 (무손실 디코딩)
    Returns: (decoded_text, detected_encoding)
    """
    if not isinstance(raw_bytes, bytes):
        return raw_bytes, hint_encoding

    # BOM 감지
    if raw_bytes[:3] == b'\xef\xbb\xbf':
        try:
            return raw_bytes.decode("utf-8-sig"), "utf-8-sig"
        except UnicodeDecodeError:
            pass

    # hint 먼저 시도
    try:
        return raw_bytes.decode(hint_encoding), hint_encoding
    except (UnicodeDecodeError, LookupError):
        pass

    # 후보 인코딩 순서대로 시도
    for enc in _ENCODING_CANDIDATES:
        if enc == hint_encoding:
            continue
        try:
            return raw_bytes.decode(enc), enc
        except (UnicodeDecodeError, LookupError):
            continue

    # 최종 fallback
    return raw_bytes.decode("latin-1"), "latin-1"


def preview_file(file_content, import_type="csv", encoding="utf-8",
                 delimiter=",", skip_header=True, max_rows=100):
    """파일 미리보기 — 첫 N행과 컬럼 정보 반환"""
    try:
        text, detected = _detect_and_decode(file_content, encoding)

        if import_type == "csv":
            reader = csv.reader(io.StringIO(text), delimiter=delimiter)
            rows = list(reader)
            if not rows:
                return {"columns": [], "rows": [], "totalRows": 0, "detectedEncoding": detected}
            columns = rows[0] if skip_header else [f"col_{i}" for i in range(len(rows[0]))]
            data_rows = rows[1:max_rows + 1] if skip_header else rows[:max_rows]
            total = len(rows) - (1 if skip_header else 0)
            return {
                "columns": columns,
                "rows": [dict(zip(columns, r)) for r in data_rows],
                "totalRows": total,
                "detectedEncoding": detected,
            }

        elif import_type == "json":
            data = json.loads(text)
            if isinstance(data, list):
                items = data
            elif isinstance(data, dict) and "data" in data:
                items = data["data"]
            else:
                items = [data]
            if not items:
                return {"columns": [], "rows": [], "totalRows": 0, "detectedEncoding": detected}
            columns = list(items[0].keys()) if isinstance(items[0], dict) else []
            return {
                "columns": columns,
                "rows": items[:max_rows],
                "totalRows": len(items),
                "detectedEncoding": detected,
            }

        return {"columns": [], "rows": [], "totalRows": 0, "error": f"Unsupported type: {import_type}"}

    except Exception as e:
        logger.error(f"Preview failed: {e}")
        return {"columns": [], "rows": [], "totalRows": 0, "error": str(e)}


def preview_files(file_list):
    """다중 파일 미리보기 — 파일 목록과 메타정보 반환"""
    items = []
    for f in file_list:
        name = f.get("name", "")
        size = f.get("size", 0)
        mime = mimetypes.guess_type(name)[0] or "application/octet-stream"
        category = _classify_file(name, mime)
        items.append({
            "name": name,
            "size": size,
            "sizeStr": _fmt_size(size),
            "mimeType": mime,
            "category": category,
        })
    return {
        "columns": ["name", "size", "mimeType", "category"],
        "rows": items,
        "totalRows": len(items),
        "isFileMode": True,
    }


def _classify_file(filename, mime_type):
    """파일 유형 분류"""
    ext = os.path.splitext(filename)[1].lower()
    if mime_type.startswith("image/"):
        return "image"
    if mime_type.startswith("video/"):
        return "video"
    if mime_type.startswith("audio/"):
        return "audio"
    if mime_type.startswith("text/") or ext in (".csv", ".json", ".xml", ".yaml", ".yml", ".log", ".md"):
        return "text"
    if ext in (".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".hwp"):
        return "document"
    if ext in (".zip", ".tar", ".gz", ".7z", ".rar"):
        return "archive"
    return "other"


def _fmt_size(size):
    if size < 1024:
        return f"{size}B"
    if size < 1048576:
        return f"{size/1024:.1f}KB"
    return f"{size/1048576:.1f}MB"


def _execute_import_files(collector, file_data_list, db_session):
    """파일 가져오기 — 다중 파일을 MinIO에 업로드하고 파일별 카탈로그 등록"""
    from backend.services.minio_client import get_minio_client
    from backend.models.catalog import DataCatalog

    cid = collector.id
    bucket = collector.target_bucket or "sdl-files"
    prefix = f"import/{cid}/"

    try:
        client = get_minio_client(db_session)
        total = len(file_data_list)
        collector.total_rows = total
        collector.status = "running"
        db_session.commit()

        imported = 0
        errors = 0
        ok_names = []  # MinIO 업로드 성공 파일명 — 호출자가 소스 정리에 사용

        for i, fdata in enumerate(file_data_list):
            try:
                fname = fdata["name"]
                content = fdata["content"]
                obj_name = prefix + fname
                mime = mimetypes.guess_type(fname)[0] or "application/octet-stream"

                client.put_object(
                    bucket, obj_name,
                    io.BytesIO(content), len(content),
                    content_type=mime,
                )

                # ── 파이프라인 라우팅: file_meta envelope 을 MQTT 로 발행 ──
                # 파이프라인이 sdl/raw/import/{cid}/# 토픽을 구독 중이면 external_file_sink 로 흘러감
                _publish_file_meta(
                    connector_type="import",
                    connector_id=cid,
                    rel_path=fname,
                    bucket=bucket,
                    object_name=obj_name,
                    file_size=len(content),
                    mime_type=mime,
                )

                imported += 1
                ok_names.append(fname)
            except Exception as e:
                errors += 1
                logger.warning(f"Import file '{fdata.get('name','')}' error: {e}")

            collector.imported_rows = imported
            collector.error_rows = errors
            collector.progress = int((i + 1) / total * 100) if total else 100
            db_session.commit()

        collector.status = "completed"
        collector.progress = 100
        collector.last_imported_at = datetime.utcnow()
        db_session.commit()
        logger.info(f"Import #{cid} files completed: {imported}/{total}")

        # 그룹 카탈로그 1건만 등록 (파일 브라우저로 탐색)
        existing = db_session.query(DataCatalog).filter_by(
            connector_type="import", connector_id=cid, tag_name=""
        ).first()
        if not existing:
            cat = DataCatalog(
                name=f"Import {collector.name} — 전체 파일",
                tag_name="",
                connector_type="import",
                connector_id=cid,
                connector_description=collector.description or "",
                description=f"데이터 가져오기 '{collector.name}'에서 수집된 {imported}개 파일",
                owner="시스템 자동",
                category="파일",
                data_level="user_created",
                sensitivity="internal",
                access_url=f"s3://{bucket}/{prefix}",
                format="mixed",
                sink_type="internal_file_sink",
                is_published=True,
            )
            db_session.add(cat)
            db_session.flush()
            _add_search_tags(db_session, cat.id, ["import", collector.name, "파일"])
        else:
            existing.description = f"데이터 가져오기 '{collector.name}'에서 수집된 {imported}개 파일"
            existing.access_url = f"s3://{bucket}/{prefix}"
        db_session.commit()
        return ok_names

    except Exception as e:
        logger.error(f"Import #{cid} files failed: {e}")
        collector.status = "error"
        collector.last_error = str(e)
        collector.error_count += 1
        db_session.commit()
        return []


def _fix_zip_filename(info):
    """ZIP 파일명을 cp437 mojibake 에서 자동 복원.

    Windows 표준 압축 도구로 만든 한국어 zip 은 일반적으로 UTF-8 플래그
    (general purpose bit 11) 를 설정하지 않고 파일명을 cp949 로 인코딩한다.
    zipfile 모듈은 그런 경우 기본 cp437 로 디코드하므로 '최대' 같은 한국어가
    '├╓┤δ' 같은 mojibake 로 보인다. 본 함수는 그 mojibake 를 cp949 → utf-8
    순으로 재해석해 복원한다. 모두 실패하면 원본 이름을 유지한다.
    """
    name = info.filename
    if info.flag_bits & 0x800:
        return name  # 압축 시 UTF-8 플래그가 명시되어 있으면 그대로 신뢰
    try:
        raw = name.encode('cp437')
    except UnicodeEncodeError:
        # cp437 로 다시 encode 가 안 되면 이미 다른 경로로 들어온 정상 문자열
        return name
    for enc in ('cp949', 'utf-8'):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return name


def extract_zip(zip_content):
    """ZIP 아카이브에서 파일 목록 추출 — 디렉토리 구조 유지 + 한국어 파일명 복원."""
    result = []
    # ZIP 내 공통 prefix 제거 (예: sample_timeseries/ → 내부 상대 경로만)
    with zipfile.ZipFile(io.BytesIO(zip_content)) as zf:
        infos = [i for i in zf.infolist() if not i.is_dir()]
        fixed_names = [_fix_zip_filename(i) for i in infos]

        # 공통 prefix 찾기 — 보정된 이름 기준
        common = ""
        if fixed_names:
            parts = fixed_names[0].split("/")
            if len(parts) > 1:
                candidate = parts[0] + "/"
                if all(n.startswith(candidate) for n in fixed_names):
                    common = candidate

        for info, name in zip(infos, fixed_names):
            if name.startswith("__MACOSX") or name.startswith(".") or "/.DS_Store" in name:
                continue
            # 공통 prefix 제거 후 상대 경로 유지
            rel_path = name[len(common):] if common and name.startswith(common) else name
            if not rel_path:
                continue
            # zf.read 는 ZipInfo 객체로 호출 — 내부 lookup 은 원본 filename 으로
            result.append({
                "name": rel_path,  # 상대 경로 유지 (한국어 복원 후)
                "path": name,
                "content": zf.read(info),
                "size": info.file_size,
            })
    return result


def scan_local_path(local_path, file_patterns=None, recursive=True, max_preview=500,
                    exclude_dirs=None):
    """서버 로컬 경로를 스캔하여 파일 목록 반환

    Args:
        local_path: 서버 로컬 디렉토리 경로
        file_patterns: 파일 패턴 목록 ["*.csv", "*.jpg"]
        recursive: 하위 디렉토리 포함 여부
        max_preview: 미리보기 최대 파일 수
        exclude_dirs: walk 에서 제외할 디렉토리명 (예: archive_subdir '.imported')
                      — 이미 가져온 파일을 재스캔·재import 하지 않도록.
    Returns:
        {"files": [...], "totalFiles": N, "totalSize": N}
    """
    import fnmatch

    if not os.path.isdir(local_path):
        return {"files": [], "totalFiles": 0, "totalSize": 0,
                "error": f"경로를 찾을 수 없습니다: {local_path}"}

    patterns = file_patterns or ["*"]
    exclude = {d for d in (exclude_dirs or []) if d}
    files = []
    total_size = 0

    if recursive:
        walker = os.walk(local_path)
    else:
        # 현재 디렉토리만
        try:
            entries = os.listdir(local_path)
        except PermissionError:
            return {"files": [], "totalFiles": 0, "totalSize": 0,
                    "error": f"접근 권한이 없습니다: {local_path}"}
        walker = [(local_path, [], [e for e in entries if os.path.isfile(os.path.join(local_path, e))])]

    for dirpath, dirnames, filenames in walker:
        # archive 영역 등 제외 디렉토리는 하위 walk 진입 자체를 차단
        if exclude:
            dirnames[:] = [d for d in dirnames if d not in exclude]
        for fname in filenames:
            # 숨김 파일 제외
            if fname.startswith("."):
                continue
            # 패턴 매칭
            matched = any(fnmatch.fnmatch(fname.lower(), p.lower()) for p in patterns)
            if not matched:
                continue

            full_path = os.path.join(dirpath, fname)
            try:
                stat = os.stat(full_path)
                rel_path = os.path.relpath(full_path, local_path)
                mime = mimetypes.guess_type(fname)[0] or "application/octet-stream"
                files.append({
                    "name": fname,
                    "path": rel_path,
                    "fullPath": full_path,
                    "size": stat.st_size,
                    "sizeStr": _fmt_size(stat.st_size),
                    "mimeType": mime,
                    "category": _classify_file(fname, mime),
                    "modifiedAt": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                })
                total_size += stat.st_size
            except (OSError, PermissionError):
                continue

    # 수정일 기준 정렬
    files.sort(key=lambda f: f.get("modifiedAt", ""), reverse=True)

    return {
        "files": files[:max_preview],
        "totalFiles": len(files),
        "totalSize": total_size,
        "totalSizeStr": _fmt_size(total_size),
    }


def start_import_from_path(collector_id):
    """서버 로컬 경로에서 파일을 읽어 Import 실행"""
    from backend.database import SessionLocal
    from backend.models.collector import ImportCollector

    def _run():
        db = SessionLocal()
        try:
            c = db.query(ImportCollector).get(collector_id)
            if not c:
                logger.error(f"Import #{collector_id} not found")
                return

            local_path = c.local_path
            patterns = c.file_patterns or ["*"]
            is_recursive = c.recursive if c.recursive is not None else True
            # archive 영역은 재스캔 대상에서 제외 — 이미 가져온 파일 재import 방지
            archive_subdir = c.archive_subdir or ".imported"

            scan = scan_local_path(local_path, patterns, is_recursive, max_preview=999999,
                                   exclude_dirs=[archive_subdir])
            if scan.get("error"):
                c.status = "error"
                c.last_error = scan["error"]
                db.commit()
                return

            all_files = scan["files"]
            if not all_files:
                c.status = "error"
                c.last_error = "스캔된 파일이 없습니다."
                db.commit()
                return

            if c.import_type == "files":
                # 비정형 파일 → MinIO 직접 업로드
                file_data_list = []
                path_map = {}  # rel_path → fullPath (소스 정리용)
                for f in all_files:
                    try:
                        with open(f["fullPath"], "rb") as fh:
                            file_data_list.append({
                                "name": f["path"],  # 상대 경로 유지
                                "content": fh.read(),
                                "size": f["size"],
                            })
                        path_map[f["path"]] = f["fullPath"]
                    except (OSError, PermissionError) as e:
                        logger.warning(f"Cannot read {f['fullPath']}: {e}")

                ok_names = _execute_import_files(c, file_data_list, db)
                if c.publish_mqtt:
                    _execute_files_mqtt_publish(c, file_data_list, db)
                # MinIO 정본화: MinIO 업로드 성공한 파일의 NFS 소스 정리
                for name in (ok_names or []):
                    fp = path_map.get(name)
                    if fp:
                        _apply_post_import_action(c, fp, name)
            else:
                # 정형 데이터 (CSV/JSON) — 파일별 순차 처리.
                # 파일 단위로 import → MQTT 발행 → 소스 정리 순서 (정리 후엔 read 불가).
                c.status = "running"
                c.total_rows = 0
                c.imported_rows = 0
                c.error_rows = 0
                c.progress = 0
                db.commit()

                total_files = len(all_files)
                for fi, f in enumerate(all_files):
                    ok = False
                    try:
                        if c.target_type == "file":
                            # target=file은 raw bytes를 그대로 MinIO에 저장하므로
                            # 메모리 풀로드 없이 파일 경로만 넘겨 스트리밍 업로드.
                            # source_filename=상대경로 → MinIO 객체 키에 디렉토리 구조 보존.
                            _execute_import_direct(c, None, db, source_filename=f["path"], accumulate=True, file_path=f["fullPath"])
                        else:
                            with open(f["fullPath"], "rb") as fh:
                                content = fh.read()
                            _execute_import_direct(c, content, db, source_filename=f["path"], accumulate=True)
                        ok = True
                    except Exception as e:
                        logger.warning(f"Import file {f['path']} error: {e}")
                        c.error_rows = (c.error_rows or 0) + 1

                    # MQTT 발행은 소스 정리 전에 (정리되면 파일 read 불가)
                    if ok and c.publish_mqtt:
                        try:
                            with open(f["fullPath"], "rb") as fh:
                                _execute_import_mqtt(c, fh.read(), db)
                        except Exception:
                            pass
                    # MinIO 정본화: 성공 파일의 NFS 소스 정리
                    if ok:
                        _apply_post_import_action(c, f["fullPath"], f["path"])

                    c.progress = int((fi + 1) / total_files * 100)
                    db.commit()

                c.status = "completed"
                c.progress = 100
                c.last_imported_at = datetime.utcnow()
                db.commit()

                _register_catalog(db, c)
        except Exception as e:
            logger.error(f"Import from path #{collector_id} error: {e}")
            try:
                c.status = "error"
                c.last_error = str(e)
                db.commit()
            except Exception:
                pass
        finally:
            db.close()
            with _import_lock:
                _import_threads.pop(collector_id, None)

    with _import_lock:
        if collector_id in _import_threads:
            raise RuntimeError(f"Import #{collector_id} already running")
        t = threading.Thread(target=_run, daemon=True, name=f"import-path-{collector_id}")
        _import_threads[collector_id] = t
        t.start()


def _parse_csv(file_content, encoding, delimiter, skip_header):
    """CSV 파일을 dict 리스트로 변환"""
    text, _ = _detect_and_decode(file_content, encoding)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        return []
    if skip_header:
        headers = rows[0]
        return [dict(zip(headers, r)) for r in rows[1:]]
    else:
        headers = [f"col_{i}" for i in range(len(rows[0]))]
        return [dict(zip(headers, r)) for r in rows]


def _parse_json(file_content, encoding):
    """JSON 파일을 dict 리스트로 변환"""
    text, _ = _detect_and_decode(file_content, encoding)
    data = json.loads(text)
    if isinstance(data, list):
        return data
    elif isinstance(data, dict) and "data" in data:
        return data["data"]
    return [data]


def _parse_xlsx(file_content, sheet_name=None, header_row=1):
    """xlsx(엑셀) 파일을 dict 리스트로 변환.

    sheet_name: 빈 값/None 이면 첫 시트.
    header_row: 1-base. 헤더가 있는 행 번호. 그 다음 행부터 데이터로 간주.
    """
    from io import BytesIO
    import openpyxl

    if isinstance(file_content, str):
        file_content = file_content.encode("utf-8")
    wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        hr = max(1, int(header_row or 1))
        headers = None
        records = []
        for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if ri < hr:
                continue
            if ri == hr:
                headers = [(str(c).strip() if c is not None else f"col_{i}") for i, c in enumerate(row)]
                continue
            if headers is None:
                continue
            # 빈 행 스킵
            if all(c is None for c in row):
                continue
            rec = {}
            for i, h in enumerate(headers):
                v = row[i] if i < len(row) else None
                rec[h] = v
            records.append(rec)
        return records
    finally:
        wb.close()


def _infer_data_type(value):
    """값으로부터 데이터 타입 추론"""
    if value is None or value == "":
        return "string", None
    if isinstance(value, (int, float)):
        return "float", float(value)
    try:
        return "float", float(value)
    except (ValueError, TypeError):
        return "string", str(value)


def _execute_import_mqtt(collector, file_content, db_session):
    """MQTT 발행 전용 — 직접 저장 후 파이프라인 연계용으로 MQTT 발행

    직접 저장이 먼저 완료된 상태에서 호출됨.
    통계(imported_rows 등)는 건드리지 않고 MQTT 발행만 수행.
    파이프라인이 구독하지 않아도 원본은 이미 저장되어 있으므로 안전.
    """
    from backend.services import mqtt_manager

    cid = collector.id
    import_type = collector.import_type
    encoding = collector.encoding or "utf-8"
    delimiter = collector.delimiter or ","

    try:
        if import_type == "csv":
            records = _parse_csv(file_content, encoding, delimiter, collector.skip_header)
        elif import_type == "json":
            records = _parse_json(file_content, encoding)
        else:
            return

        ts_col = collector.timestamp_column
        tag_col = collector.tag_column
        value_cols = collector.value_columns or []
        col_map = collector.column_mapping or {}
        published = 0

        for record in records:
            try:
                if value_cols:
                    for vc in value_cols:
                        raw_val = record.get(vc, "")
                        tag_name = col_map.get(vc, vc)
                        dtype, val = _infer_data_type(raw_val)
                        if val is not None:
                            mqtt_manager.publish_raw(
                                connector_type="import",
                                connector_id=cid,
                                tag_name=tag_name,
                                value=val,
                                data_type=dtype,
                                quality=100,
                            )
                else:
                    skip = {ts_col, tag_col} if ts_col or tag_col else set()
                    for col, raw_val in record.items():
                        if col in skip:
                            continue
                        tag_name = col_map.get(col, col)
                        dtype, val = _infer_data_type(raw_val)
                        if val is not None:
                            mqtt_manager.publish_raw(
                                connector_type="import",
                                connector_id=cid,
                                tag_name=tag_name,
                                value=val,
                                data_type=dtype,
                                quality=100,
                            )
                published += 1
            except Exception as e:
                logger.warning(f"Import #{cid} MQTT publish error: {e}")

        logger.info(f"Import #{cid} MQTT published: {published} records for pipeline processing")

    except Exception as e:
        logger.error(f"Import #{cid} MQTT publish failed: {e}")


def _execute_files_mqtt_publish(collector, file_data_list, db_session):
    """비정형 파일 메타데이터를 MQTT로 발행 — 파이프라인에서 파일 메타 처리 가능

    각 파일의 메타정보(파일명, 크기, 유형, S3 경로)를 MQTT 메시지로 발행.
    파이프라인에서 메타데이터 기반 추가 처리(분류, 태깅, 알림 등)에 활용.
    """
    from backend.services import mqtt_manager

    cid = collector.id
    bucket = collector.target_bucket or "sdl-files"
    date_prefix = datetime.utcnow().strftime("%Y%m%d")
    published = 0

    for fdata in file_data_list:
        try:
            fname = fdata["name"]
            fsize = fdata.get("size", len(fdata.get("content", b"")))
            mime = mimetypes.guess_type(fname)[0] or "application/octet-stream"
            category = _classify_file(fname, mime)
            obj_path = f"import/{cid}/{date_prefix}/{fname}"

            # 파일 메타정보를 JSON value로 발행
            meta_value = {
                "fileName": fname,
                "fileSize": fsize,
                "mimeType": mime,
                "category": category,
                "bucket": bucket,
                "objectPath": obj_path,
                "accessUrl": f"s3://{bucket}/{obj_path}",
            }
            mqtt_manager.publish_raw(
                connector_type="import",
                connector_id=cid,
                tag_name=fname,
                value=meta_value,
                data_type="json",
                quality=100,
            )
            published += 1
        except Exception as e:
            logger.warning(f"Import #{cid} file MQTT publish '{fname}' error: {e}")

    logger.info(f"Import #{cid} file metadata MQTT published: {published} files")


def _execute_import_direct(collector, file_content, db_session, source_filename=None, accumulate=False, file_path=None):
    """직접 저장 모드 — DB INSERT 또는 MinIO 업로드.

    source_filename: 다중 파일 반복 처리(예: 서버 경로 모드) 시 각 파일의 원본 이름을 전달.
    target_type='file' 일 때 MinIO 객체 키에 그대로 사용되어 파일별 객체로 분리 저장됨.
    None이면 collector.file_name → 최후 fallback 'import_{cid}.dat' 순으로 결정.

    accumulate: True 이면 imported_rows/total_rows/error_rows 를 덮어쓰지 않고 누적,
    상태(status)/진행률(progress)/last_imported_at 도 호출자가 설정하도록 위임한다.
    서버 경로 모드처럼 N개 파일을 순차로 호출할 때 사용.

    file_path: target_type='file' 한정. 주어지면 file_content 대신 디스크에서 직접
    스트리밍 업로드한다(메모리 풀로드·CSV 파싱 모두 생략). GB급 서버 경로 모드용.
    """
    from backend.models.storage import TsdbConfig, RdbmsConfig
    from sqlalchemy import create_engine, text

    cid = collector.id
    import_type = collector.import_type
    encoding = collector.encoding or "utf-8"
    delimiter = collector.delimiter or ","
    target_type = collector.target_type
    batch_size = collector.batch_size or 1000

    try:
        # target=file은 raw 바이트를 그대로 MinIO에 저장 — CSV/JSON 파싱 불필요.
        # 파싱은 GB급 파일에서 메모리 폭증 + 무의미한 CPU 사용의 주범이었음.
        if target_type == "file":
            from backend.services.minio_client import get_minio_client
            client = get_minio_client(db_session)
            bucket = collector.target_bucket or "sdl-files"
            # 결정적 객체 키 — 날짜는 '소스 파일 mtime' 기준 (실행일 아님).
            # 같은 파일을 재import 하면 같은 키 → 덮어쓰기 → 중복 객체 누적 방지.
            if file_path and os.path.exists(file_path):
                date_str = datetime.utcfromtimestamp(
                    os.path.getmtime(file_path)).strftime("%Y%m%d")
            else:
                date_str = datetime.utcnow().strftime("%Y%m%d")
            prefix = f"import/{cid}/{date_str}/"
            obj_name = prefix + (source_filename or collector.file_name or f"import_{cid}.dat")

            if file_path:
                # 디스크에서 직접 스트리밍 업로드(multipart). 풀로드 없음.
                size = os.path.getsize(file_path)
                with open(file_path, "rb") as fh:
                    client.put_object(bucket, obj_name, fh, size)
                upload_size = size
            else:
                if isinstance(file_content, str):
                    file_content = file_content.encode(encoding)
                from io import BytesIO
                client.put_object(bucket, obj_name, BytesIO(file_content), len(file_content))
                upload_size = len(file_content)

            # ── 파이프라인 라우팅: file_meta envelope 발행 ──
            _publish_file_meta(
                connector_type="import",
                connector_id=cid,
                rel_path=source_filename or collector.file_name or os.path.basename(obj_name),
                bucket=bucket,
                object_name=obj_name,
                file_size=upload_size,
                mime_type=mimetypes.guess_type(obj_name)[0] or "application/octet-stream",
            )

            if accumulate:
                collector.imported_rows = (collector.imported_rows or 0) + 1
                collector.total_rows = (collector.total_rows or 0) + 1
            else:
                collector.imported_rows = 1
                collector.total_rows = 1
                collector.error_rows = 0
                collector.status = "completed"
                collector.progress = 100
                collector.last_imported_at = datetime.utcnow()
            db_session.commit()
            logger.info(f"Import #{cid} file uploaded: {bucket}/{obj_name}")
            return

        if import_type == "csv":
            records = _parse_csv(file_content, encoding, delimiter, collector.skip_header)
        elif import_type == "json":
            records = _parse_json(file_content, encoding)
        elif import_type == "xlsx":
            records = _parse_xlsx(file_content, collector.sheet_name, collector.header_row)
        else:
            raise ValueError(f"Unsupported import type: {import_type}")

        total = len(records)
        if not accumulate:
            collector.total_rows = total
            collector.status = "running"
        db_session.commit()

        col_map = collector.column_mapping or {}
        ts_col = collector.timestamp_column

        if target_type == "tsdb":
            # TSDB 직접 저장 — time_series_data 테이블에 INSERT
            # measurement가 '*'이면 각 컬럼명을 measurement로 사용
            measurement_cfg = collector.target_measurement or collector.name
            auto_measurement = (measurement_cfg.strip() == "*")
            tsdb_id = collector.target_id or 0
            value_cols = collector.value_columns or []
            imported = 0
            errors = 0

            for i, record in enumerate(records):
                try:
                    ts_val = record.get(ts_col, "") if ts_col else None
                    try:
                        ts_dt = datetime.fromisoformat(str(ts_val).replace("Z", "+00:00")) if ts_val else datetime.utcnow()
                    except (ValueError, TypeError):
                        ts_dt = datetime.utcnow()

                    cols_to_insert = value_cols if value_cols else [k for k in record if k != ts_col]
                    for vc in cols_to_insert:
                        raw_val = record.get(vc, "")
                        tag_name = col_map.get(vc, vc)
                        dtype, val = _infer_data_type(raw_val)
                        if val is None:
                            continue
                        meas = tag_name if auto_measurement else measurement_cfg
                        db_session.execute(text("""
                            INSERT INTO time_series_data
                            (tsdb_id, tag_name, connector_type, connector_id, measurement,
                             value, value_str, data_type, unit, quality, timestamp, created_at)
                            VALUES (:tsdb_id, :tag, :ctype, :cid, :meas,
                                    :val, :vstr, :dtype, :unit, :qual, :ts, :now)
                        """), {
                            "tsdb_id": tsdb_id,
                            "tag": tag_name, "ctype": "import", "cid": cid,
                            "meas": meas,
                            "val": val if dtype == "float" else None,
                            "vstr": str(raw_val),
                            "dtype": dtype, "unit": "", "qual": 100,
                            "ts": ts_dt, "now": datetime.utcnow(),
                        })
                    imported += 1
                except Exception as e:
                    errors += 1
                    logger.warning(f"Import TSDB row {i} error: {e}")

                if (i + 1) % batch_size == 0:
                    db_session.commit()
                    collector.imported_rows = imported
                    collector.progress = int((i + 1) / total * 100) if total else 100
                    db_session.commit()

            db_session.commit()
            if accumulate:
                collector.imported_rows = (collector.imported_rows or 0) + imported
                collector.error_rows = (collector.error_rows or 0) + errors
                collector.total_rows = (collector.total_rows or 0) + total
            else:
                collector.imported_rows = imported
                collector.error_rows = errors

            # TagMetadata 일괄 생성 (직접 저장 모드용)
            _ensure_tag_metadata(db_session, "import", cid, collector.name, records, ts_col, col_map, value_cols)

        elif target_type == "rdbms":
            # RDBMS 직접 저장 — 외부 DB에 INSERT (테이블 자동 생성)
            cfg = db_session.query(RdbmsConfig).get(collector.target_id)
            if not cfg:
                raise ValueError(f"RDBMS config #{collector.target_id} not found")

            engine = create_engine(_build_rdbms_url(cfg), pool_pre_ping=True)
            table_name = collector.target_table or f"import_{cid}"
            imported = 0
            errors = 0

            with engine.connect() as conn:
                # 테이블 자동 생성 (첫 번째 레코드 기준)
                if records:
                    first = records[0]
                    mapped_first = {}
                    for k, v in first.items():
                        mk = col_map.get(k, k).strip()
                        if not mk:
                            continue  # 빈 컬럼명 건너뛰기
                        mapped_first[mk] = v
                    col_defs = ", ".join(
                        f'"{c}" TEXT' for c in mapped_first.keys()
                    )
                    conn.execute(text(
                        f'CREATE TABLE IF NOT EXISTS "{table_name}" '
                        f'(id SERIAL PRIMARY KEY, {col_defs}, imported_at TIMESTAMP DEFAULT NOW())'
                    ))
                    conn.commit()
                    logger.info(f"Import #{cid}: table '{table_name}' ensured ({len(mapped_first)} columns)")

                for i, record in enumerate(records):
                    try:
                        mapped = {}
                        for k, v in record.items():
                            mapped_key = col_map.get(k, k).strip()
                            if not mapped_key:
                                continue  # 빈 컬럼명 건너뛰기
                            mapped[mapped_key] = v

                        cols = ", ".join(f'"{c}"' for c in mapped.keys())
                        placeholders = ", ".join(f":{k}" for k in mapped.keys())
                        conn.execute(
                            text(f'INSERT INTO "{table_name}" ({cols}) VALUES ({placeholders})'),
                            mapped
                        )
                        imported += 1
                    except Exception as e:
                        errors += 1
                        logger.warning(f"Import RDBMS row {i} error: {e}")

                    if (i + 1) % batch_size == 0:
                        conn.commit()
                        collector.imported_rows = imported
                        collector.progress = int((i + 1) / total * 100) if total else 100
                        db_session.commit()

                conn.commit()
            engine.dispose()
            if accumulate:
                collector.imported_rows = (collector.imported_rows or 0) + imported
                collector.error_rows = (collector.error_rows or 0) + errors
                collector.total_rows = (collector.total_rows or 0) + total
            else:
                collector.imported_rows = imported
                collector.error_rows = errors

        if not accumulate:
            collector.status = "completed"
            collector.progress = 100
            collector.last_imported_at = datetime.utcnow()
        db_session.commit()
        logger.info(f"Import #{cid} direct file done: {collector.imported_rows}/{collector.total_rows} rows")

    except Exception as e:
        logger.error(f"Import #{cid} direct failed: {e}")
        collector.status = "error"
        collector.last_error = str(e)
        collector.error_count += 1
        db_session.commit()


def _concat_text_files(files, drop_header_after_first=True):
    """CSV/TSV/TXT 다중 파일을 바이트 단위로 연결.

    drop_header_after_first=True 인 경우, 두번째 파일부터 첫 줄(헤더 추정) 제거.
    """
    if not files:
        return b""
    parts = [files[0]["content"]]
    if not drop_header_after_first:
        parts.extend(f["content"] for f in files[1:])
    else:
        for f in files[1:]:
            c = f["content"]
            nl = c.find(b"\n")
            if nl >= 0:
                parts.append(c[nl+1:])
            else:
                continue
    return b"\n".join(parts)


def _concat_json_files(files):
    """JSON 다중 파일: 각 파일을 파싱해 배열을 누적."""
    items = []
    for f in files:
        try:
            txt = f["content"].decode("utf-8", errors="replace")
            data = json.loads(txt)
            if isinstance(data, list):
                items.extend(data)
            elif isinstance(data, dict):
                items.append(data)
        except Exception as e:
            logger.warning(f"JSON parse skipped for {f.get('name','?')}: {e}")
    return json.dumps(items, ensure_ascii=False).encode("utf-8")


def start_import(collector_id, file_content, file_data_list=None):
    """비동기 Import 실행 — 백그라운드 스레드에서 처리

    Args:
        collector_id: ImportCollector ID
        file_content: 단일 파일 바이트 (CSV/JSON용)
        file_data_list: 다중 파일 리스트 (파일 가져오기용) [{"name": ..., "content": ...}, ...]
    """
    from backend.database import SessionLocal
    from backend.models.collector import ImportCollector

    def _run():
        db = SessionLocal()
        try:
            c = db.query(ImportCollector).get(collector_id)
            if not c:
                logger.error(f"Import collector #{collector_id} not found")
                return

            if c.import_type == "files":
                # 비정형 파일 가져오기
                files = file_data_list or []
                if not files and file_content:
                    fname = c.file_name or f"import_{collector_id}"
                    if fname.lower().endswith(".zip"):
                        files = extract_zip(file_content)
                    else:
                        files = [{"name": fname, "content": file_content, "size": len(file_content)}]
                # ① 항상 MinIO에 저장 (원본 보장 + 파일별 카탈로그)
                _execute_import_files(c, files, db)
                # ② 파이프라인 연계 시 파일 메타정보를 MQTT 발행
                if c.publish_mqtt:
                    _execute_files_mqtt_publish(c, files, db)
            else:
                # 정형 데이터 (CSV/JSON) 가져오기
                multi_files = bool(file_data_list and len(file_data_list) > 0)

                if c.target_type == "file" and multi_files and len(file_data_list) > 1:
                    # target=file + 다중 파일: 병합하지 않고 원본 N개를 그대로 N개 객체로 업로드
                    # (병합 시 collector.file_name="N files" 가 그대로 MinIO 객체 키로 사용되는 회귀 방지.
                    #  서버 경로 모드와 동일한 파일별 적재 동작으로 통일.)
                    c.status = "running"
                    c.total_rows = 0
                    c.imported_rows = 0
                    c.error_rows = 0
                    c.progress = 0
                    db.commit()
                    total_files = len(file_data_list)
                    for fi, fdata in enumerate(file_data_list):
                        try:
                            _execute_import_direct(
                                c, fdata["content"], db,
                                source_filename=fdata["name"],
                                accumulate=True,
                            )
                        except Exception as e:
                            logger.warning(f"Import file {fdata.get('name','?')} error: {e}")
                            c.error_rows = (c.error_rows or 0) + 1
                        c.progress = int((fi + 1) / total_files * 100)
                        db.commit()
                    c.status = "completed"
                    c.progress = 100
                    c.last_imported_at = datetime.utcnow()
                    db.commit()
                    _register_catalog(db, c)
                    if c.publish_mqtt:
                        for fdata in file_data_list:
                            try:
                                _execute_import_mqtt(c, fdata["content"], db)
                            except Exception:
                                pass
                else:
                    # 기존 흐름: TSDB/RDBMS 다중 파일은 한 덩어리로 적재가 자연스러움 → 병합 유지.
                    # 단일 파일이면 file_content 그대로 사용.
                    if multi_files:
                        if (c.import_type or "csv").lower() == "json":
                            merged_content = _concat_json_files(file_data_list)
                        else:
                            merged_content = _concat_text_files(file_data_list, drop_header_after_first=not c.skip_header)
                    else:
                        merged_content = file_content
                    # ① 항상 직접 저장 (원본 보장)
                    _execute_import_direct(c, merged_content, db)
                    _register_catalog(db, c)
                    # ② 파이프라인 연계 시 MQTT 추가 발행
                    if c.publish_mqtt:
                        _execute_import_mqtt(c, merged_content, db)
        except Exception as e:
            logger.error(f"Import thread #{collector_id} error: {e}")
        finally:
            db.close()
            with _import_lock:
                _import_threads.pop(collector_id, None)

    with _import_lock:
        if collector_id in _import_threads:
            raise RuntimeError(f"Import #{collector_id} already running")
        t = threading.Thread(target=_run, daemon=True, name=f"import-{collector_id}")
        _import_threads[collector_id] = t
        t.start()


def _register_catalog(db, collector):
    """Import 결과를 데이터 카탈로그에 등록 — 기존 커넥터와 동일 수준"""
    from backend.models.catalog import DataCatalog

    target_type = collector.target_type
    import_type = collector.import_type

    # target_type → sink_type 매핑
    _SINK_MAP = {"tsdb": "internal_tsdb_sink", "rdbms": "internal_rdbms_sink", "file": "internal_file_sink"}
    sink_type = _SINK_MAP.get(target_type, "")

    if target_type == "tsdb":
        access_url = f"tsdb://{collector.target_measurement or collector.name}"
        fmt = "timeseries"
        category_val = "시계열"
    elif target_type == "rdbms":
        access_url = f"rdbms://{collector.target_id}/{collector.target_table}"
        fmt = "table"
        category_val = "관계형"
    elif target_type == "file":
        access_url = f"s3://{collector.target_bucket}/import/{collector.id}/"
        fmt = import_type
        category_val = "파일"
    else:
        access_url = ""
        fmt = import_type
        category_val = "기타"

    # 컬럼 정보 수집 (schema_info)
    schema_info = ""
    if collector.value_columns:
        schema_info = ", ".join(collector.value_columns)
    elif collector.column_mapping:
        schema_info = ", ".join(collector.column_mapping.values())

    existing = db.query(DataCatalog).filter_by(
        connector_type="import", connector_id=collector.id, tag_name=""
    ).first()

    if existing:
        existing.description = collector.description
        existing.access_url = access_url
        existing.format = fmt
        existing.schema_info = schema_info
        existing.sink_type = sink_type
        existing.connector_description = collector.description or ""
        existing.updated_at = datetime.utcnow()
    else:
        cat = DataCatalog(
            name=f"Import {collector.name} — 전체 데이터",
            tag_name="",
            connector_type="import",
            connector_id=collector.id,
            connector_description=collector.description or "",
            description=f"데이터 가져오기 '{collector.name}'에서 적재된 {category_val} 데이터",
            owner="시스템 자동",
            category=category_val,
            data_level="user_created",
            sensitivity="internal",
            access_url=access_url,
            format=fmt,
            schema_info=schema_info,
            sink_type=sink_type,
            is_published=True,
        )
        db.add(cat)
        db.flush()
        # 검색 태그 생성
        tags = ["import", collector.name, category_val, import_type]
        if collector.target_measurement:
            tags.append(collector.target_measurement)
        if collector.target_table:
            tags.append(collector.target_table)
        _add_search_tags(db, cat.id, tags)

    # 태그별 카탈로그는 생성하지 않음 — TagMetadata에서 거버넌스 통합 관리
    db.commit()


def _ensure_tag_metadata(db_session, connector_type, connector_id, connector_name,
                         records, ts_col, col_map, value_cols):
    """Import 직접 저장 후 TagMetadata 일괄 생성"""
    from backend.models.metadata import TagMetadata

    # 수집된 컬럼 목록 파악
    if value_cols:
        cols = value_cols
    elif records:
        cols = [k for k in records[0].keys() if k != ts_col]
    else:
        return

    for vc in cols:
        tag_name = col_map.get(vc, vc) if col_map else vc
        if not tag_name or not tag_name.strip():
            continue

        existing = db_session.query(TagMetadata).filter_by(
            connector_type=connector_type,
            connector_id=connector_id,
            tag_name=tag_name,
        ).first()

        if not existing:
            from backend.services.metadata_tracker import _guess_category
            meta = TagMetadata(
                tag_name=tag_name,
                connector_type=connector_type,
                connector_id=connector_id,
                connector_name=connector_name or "",
                data_type="float",
                sample_count=len(records),
                first_seen_at=datetime.utcnow(),
                last_seen_at=datetime.utcnow(),
                is_active=False,
                mqtt_topic=f"sdl/raw/{connector_type}/{connector_id}/{tag_name}",
                owner="시스템 자동",
                category=_guess_category(tag_name),
                data_level="user_created",
                sensitivity="internal",
                is_published=True,
            )
            db_session.add(meta)

    db_session.commit()
    logger.info(f"TagMetadata ensured for {connector_type}#{connector_id}: {len(cols)} tags")


def _add_search_tags(db_session, catalog_id, tags):
    """카탈로그 검색 태그 생성"""
    from backend.models.catalog import CatalogSearchTag
    for tag in tags:
        if tag and tag.strip():
            existing = db_session.query(CatalogSearchTag).filter_by(
                catalog_id=catalog_id, tag=tag.strip()
            ).first()
            if not existing:
                db_session.add(CatalogSearchTag(catalog_id=catalog_id, tag=tag.strip()))
    db_session.flush()


def republish(collector_id):
    """저장된 원본 데이터를 MQTT로 재발행 — 파이프라인 재처리용

    이미 TSDB/RDBMS에 저장된 Import 데이터를 다시 읽어서
    sdl/raw/import/{id}/{tag} 토픽에 발행합니다.
    파이프라인이 구독 중이면 실시간으로 처리됩니다.
    """
    from backend.database import SessionLocal
    from backend.models.collector import ImportCollector

    def _run():
        db = SessionLocal()
        try:
            c = db.query(ImportCollector).get(collector_id)
            if not c:
                logger.error(f"Republish: Import #{collector_id} not found")
                return

            cid = c.id
            target_type = c.target_type

            if c.import_type == "files":
                # 비정형 파일 → 메타정보 재발행
                _republish_files(c, db)
            elif target_type == "tsdb":
                # TSDB 원본 → MQTT 재발행
                _republish_from_tsdb(c, db)
            elif target_type == "rdbms":
                # RDBMS 원본 → MQTT 재발행
                _republish_from_rdbms(c, db)

        except Exception as e:
            logger.error(f"Republish #{collector_id} error: {e}")
            c.status = "error"
            c.last_error = f"Republish failed: {e}"
            db.commit()
        finally:
            db.close()
            with _import_lock:
                _import_threads.pop(collector_id, None)

    with _import_lock:
        if collector_id in _import_threads:
            raise RuntimeError(f"Import #{collector_id} already running")
        t = threading.Thread(target=_run, daemon=True, name=f"republish-{collector_id}")
        _import_threads[collector_id] = t
        t.start()


def _republish_from_tsdb(collector, db_session):
    """TSDB에 저장된 원본 데이터를 MQTT로 재발행"""
    from backend.services import mqtt_manager
    from sqlalchemy import text

    cid = collector.id
    batch_size = collector.batch_size or 1000

    collector.status = "running"
    collector.progress = 0
    db_session.commit()

    rows = db_session.execute(text("""
        SELECT tag_name, value, value_str, data_type, unit, quality, timestamp
        FROM time_series_data
        WHERE connector_type = 'import' AND connector_id = :cid
        ORDER BY timestamp
    """), {"cid": cid}).fetchall()

    total = len(rows)
    collector.total_rows = total
    db_session.commit()

    published = 0
    for i, row in enumerate(rows):
        try:
            val = row.value if row.value is not None else row.value_str
            mqtt_manager.publish_raw(
                connector_type="import",
                connector_id=cid,
                tag_name=row.tag_name,
                value=val,
                data_type=row.data_type or "float",
                unit=row.unit or "",
                quality=row.quality or 100,
            )
            published += 1
        except Exception as e:
            logger.warning(f"Republish #{cid} row {i} error: {e}")

        if (i + 1) % batch_size == 0 or i == total - 1:
            collector.progress = int((i + 1) / total * 100) if total else 100
            collector.imported_rows = published
            db_session.commit()

    collector.status = "completed"
    collector.progress = 100
    collector.imported_rows = published
    collector.last_imported_at = datetime.utcnow()
    db_session.commit()
    logger.info(f"Republish #{cid} completed: {published}/{total} messages published to MQTT")


def _republish_from_rdbms(collector, db_session):
    """RDBMS에 저장된 원본 데이터를 MQTT로 재발행"""
    from backend.services import mqtt_manager
    from backend.models.storage import RdbmsConfig
    from sqlalchemy import create_engine, text

    cid = collector.id
    batch_size = collector.batch_size or 1000

    cfg = db_session.query(RdbmsConfig).get(collector.target_id)
    if not cfg:
        raise ValueError(f"RDBMS config #{collector.target_id} not found")

    collector.status = "running"
    collector.progress = 0
    db_session.commit()

    engine = create_engine(_build_rdbms_url(cfg), pool_pre_ping=True)
    table_name = collector.target_table
    col_map = collector.column_mapping or {}

    with engine.connect() as conn:
        rows = conn.execute(text(f"SELECT * FROM {table_name}")).fetchall()
        columns = rows[0]._fields if rows else []

    total = len(rows)
    collector.total_rows = total
    db_session.commit()

    published = 0
    ts_col = collector.timestamp_column
    for i, row in enumerate(rows):
        try:
            row_dict = dict(zip(columns, row))
            skip = {ts_col} if ts_col else set()
            for col, raw_val in row_dict.items():
                if col in skip:
                    continue
                tag_name = col_map.get(col, col)
                dtype, val = _infer_data_type(raw_val)
                if val is not None:
                    mqtt_manager.publish_raw(
                        connector_type="import",
                        connector_id=cid,
                        tag_name=tag_name,
                        value=val,
                        data_type=dtype,
                        quality=100,
                    )
            published += 1
        except Exception as e:
            logger.warning(f"Republish #{cid} RDBMS row {i} error: {e}")

        if (i + 1) % batch_size == 0 or i == total - 1:
            collector.progress = int((i + 1) / total * 100) if total else 100
            collector.imported_rows = published
            db_session.commit()

    engine.dispose()
    collector.status = "completed"
    collector.progress = 100
    collector.imported_rows = published
    collector.last_imported_at = datetime.utcnow()
    db_session.commit()
    logger.info(f"Republish #{cid} RDBMS completed: {published}/{total} rows → MQTT")


def _republish_files(collector, db_session):
    """MinIO에 저장된 파일의 메타정보를 MQTT로 재발행"""
    from backend.services import mqtt_manager
    from backend.services.minio_client import get_minio_client

    cid = collector.id
    bucket = collector.target_bucket or "sdl-files"
    prefix = f"import/{cid}/"

    collector.status = "running"
    collector.progress = 0
    db_session.commit()

    client = get_minio_client(db_session)
    objects = list(client.list_objects(bucket, prefix=prefix, recursive=True))
    total = len(objects)
    collector.total_rows = total
    db_session.commit()

    published = 0
    for i, obj in enumerate(objects):
        try:
            fname = os.path.basename(obj.object_name)
            mime = mimetypes.guess_type(fname)[0] or "application/octet-stream"
            category = _classify_file(fname, mime)

            meta_value = {
                "fileName": fname,
                "fileSize": obj.size,
                "mimeType": mime,
                "category": category,
                "bucket": bucket,
                "objectPath": obj.object_name,
                "accessUrl": f"s3://{bucket}/{obj.object_name}",
            }
            mqtt_manager.publish_raw(
                connector_type="import",
                connector_id=cid,
                tag_name=fname,
                value=meta_value,
                data_type="json",
                quality=100,
            )
            published += 1
        except Exception as e:
            logger.warning(f"Republish #{cid} file '{obj.object_name}' error: {e}")

        collector.progress = int((i + 1) / total * 100) if total else 100
        collector.imported_rows = published
        db_session.commit()

    collector.status = "completed"
    collector.progress = 100
    collector.imported_rows = published
    collector.last_imported_at = datetime.utcnow()
    db_session.commit()
    logger.info(f"Republish #{cid} files completed: {published}/{total} file metadata → MQTT")


def is_running(collector_id):
    """Import 실행 중 여부 확인"""
    with _import_lock:
        return collector_id in _import_threads
