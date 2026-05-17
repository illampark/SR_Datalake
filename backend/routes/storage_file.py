import logging
import os
import io
import shutil
import threading
import time
import psutil
from datetime import datetime
from flask import Blueprint, request, jsonify, send_file
from minio.error import S3Error
from backend.database import SessionLocal
from backend.models.storage import FileCleanupPolicy
from backend.config import MINIO_BUCKETS
from backend.services.audit_logger import audit_route
from backend.services.minio_client import get_minio_client, get_minio_config

logger = logging.getLogger(__name__)


import json as _json   # cleanup 진행 파일(_cleanup_progress_*) 직렬화용

file_bp = Blueprint("storage_file", __name__, url_prefix="/api/storage/file")


def _ok(data=None, meta=None):
    resp = {"success": True, "data": data, "error": None}
    if meta:
        resp["meta"] = meta
    return jsonify(resp)


def _err(msg, code="ERROR", status=400):
    return jsonify({"success": False, "data": None, "error": {"code": code, "message": msg}}), status


def _db():
    return SessionLocal()


def _get_minio():
    db = SessionLocal()
    try:
        return get_minio_client(db)
    finally:
        db.close()


def _fmt_bytes(b):
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if abs(b) < 1024:
            return f"{b:.1f} {unit}"
        b /= 1024
    return f"{b:.1f} PB"


_DISK_PATH_CANDIDATES = (
    "/app/static/uploads",   # bind mount target (DATA_ROOT/sdl-uploads)
    "/data",                 # host data dir (스테이징 등)
    "/",                     # 최후 fallback
)


def _resolve_disk_path():
    for p in _DISK_PATH_CANDIDATES:
        if os.path.exists(p):
            return p
    return "/"


FILE_TYPE_MAP = {
    ".log": "log", ".csv": "csv", ".tsv": "csv",
    ".json": "json", ".jsonl": "json", ".ndjson": "json",
    ".png": "image", ".jpg": "image", ".jpeg": "image",
    ".gif": "image", ".bmp": "image", ".svg": "image",
    ".pdf": "doc", ".doc": "doc", ".docx": "doc",
    ".xls": "excel", ".xlsx": "excel",
    ".ppt": "doc", ".pptx": "doc", ".txt": "doc",
    ".tar": "backup", ".gz": "backup", ".zip": "backup",
    ".bak": "backup", ".7z": "backup", ".rar": "backup",
}

FILE_TYPE_LABELS = {
    "log": "로그 파일",
    "csv": "CSV 파일",
    "json": "JSON / JSONL",
    "image": "이미지",
    "doc": "문서",
    "excel": "Excel",
    "backup": "백업 파일",
    "other": "기타",
}


def _classify(filename):
    ext = os.path.splitext(filename)[1].lower()
    return FILE_TYPE_MAP.get(ext, "other"), ext


# ──────────────────────────────────────────────
# STR-011: GET /api/storage/file/status
# ──────────────────────────────────────────────
@file_bp.route("/status", methods=["GET"])
def get_storage_status():
    db = SessionLocal()
    cfg = get_minio_config(db)
    try:
        # ── 1) MinIO 사용량 — minio_object 인덱스 집계 ──
        from sqlalchemy import text as _sql_text1
        agg = {r[0]: (int(r[1]), int(r[2])) for r in db.execute(_sql_text1(
            "SELECT bucket, COUNT(*), COALESCE(SUM(size),0) FROM minio_object GROUP BY bucket"
        )).fetchall()}
        minio_total_size = minio_total_objects = 0
        bucket_details = []
        for b in MINIO_BUCKETS:
            c, s = agg.get(b, (0, 0))
            minio_total_objects += c
            minio_total_size += s
            bucket_details.append({"bucket": b, "size_bytes": s,
                                   "size_display": _fmt_bytes(s), "object_count": c})
        minio_status = "connected"
        minio_error = None

        # ── 2) local_path import_collector 들의 디스크 사용량 (file_index SQL) ──
        # 이전엔 모든 collector 의 os.walk 로 stat → NFS 위에서 분 단위 timeout 위험.
        # 이제 file_indexer 가 채운 file_index 테이블에서 집계.
        from sqlalchemy import text as _sql_text2
        local_total_size = 0
        local_total_objects = 0
        local_paths = []
        try:
            from backend.models.collector import ImportCollector
            ic_rows = (
                db.query(ImportCollector)
                  .filter(ImportCollector.source_mode == "local_path")
                  .all()
            )
            sums = {}
            if ic_rows:
                ids = tuple(ic.id for ic in ic_rows)
                res = db.execute(_sql_text2(
                    "SELECT collector_id, COUNT(*), COALESCE(SUM(size),0) "
                    "FROM file_index "
                    "WHERE collector_id IN :ids AND is_dir = false "
                    "GROUP BY collector_id"
                ), {"ids": ids}).fetchall()
                sums = {r[0]: (int(r[1] or 0), int(r[2] or 0)) for r in res}
            for ic in ic_rows:
                p = ic.local_path or ""
                if not p:
                    continue
                lcount, lsize = sums.get(ic.id, (0, 0))
                local_total_size += lsize
                local_total_objects += lcount
                local_paths.append({
                    "collectorId": ic.id,
                    "collectorName": getattr(ic, "name", "") or f"import #{ic.id}",
                    "path": p,
                    "size_bytes": lsize,
                    "size_display": _fmt_bytes(lsize),
                    "object_count": lcount,
                })
        except Exception as e:
            logger.warning("local_path 사용량 집계 실패: %s", e)

        # ── 3) 디스크 전체 capacity ──
        cap_override = request.args.get("capacityGB")
        disk_path = None
        if cap_override is not None:
            try:
                total_gb = float(cap_override)
                # avail 는 합산 used 기준으로 계산 (아래에서)
            except (TypeError, ValueError):
                cap_override = None
        if cap_override is None:
            disk_path = _resolve_disk_path()
            disk = psutil.disk_usage(disk_path)
            total_gb = disk.total / (1024 ** 3)

        minio_used_gb = minio_total_size / (1024 ** 3)
        local_used_gb = local_total_size / (1024 ** 3)
        # MinIO 정본화: 데이터레이크 사용량 합계 = MinIO 전용.
        # local_path 는 'import 대기함(inbox)' 일 뿐 — 정본 저장소가 아니므로
        # totalUsedGB / totalObjects 에 합산하지 않는다 (이중 계상 방지).
        total_used_gb = minio_used_gb
        avail_gb = max(total_gb - total_used_gb, 0)

        def _pct(used):
            return round((used / total_gb) * 100, 1) if total_gb > 0 else 0

        # import 대기함(inbox) — 정보 표시용. 합계 미반영.
        inbox = {
            "usedGB": round(local_used_gb, 4),
            "sizeBytes": local_total_size,
            "sizeDisplay": _fmt_bytes(local_total_size),
            "objectCount": local_total_objects,
            "usagePercent": _pct(local_used_gb),
            "paths": local_paths,
        }

        return _ok({
            # ── 하위 호환 (기존 stat 카드용 — MinIO 정본 기준) ──
            "totalGB": round(total_gb, 2),
            "usedGB": round(total_used_gb, 4),
            "availableGB": round(avail_gb, 2),
            "usagePercent": _pct(total_used_gb),
            "totalObjects": minio_total_objects,
            "buckets": bucket_details,
            "storageType": "MinIO S3",
            "endpoint": cfg["endpoint"],
            "status": minio_status,
            "diskPath": disk_path,
            "snapshot_at": datetime.utcnow().isoformat(),
            # ── 정본 저장소 (MinIO) ──
            "minio": {
                "usedGB": round(minio_used_gb, 4),
                "sizeBytes": minio_total_size,
                "sizeDisplay": _fmt_bytes(minio_total_size),
                "objectCount": minio_total_objects,
                "usagePercent": _pct(minio_used_gb),
                "status": minio_status,
                "error": minio_error,
                "buckets": bucket_details,
            },
            # ── import 대기함 (inbox) — 합계 미반영 ──
            "inbox": inbox,
            "localPath": inbox,  # 하위 호환 별칭 (deprecated)
        })
    except Exception as e:
        return _ok({
            "totalGB": 0, "usedGB": 0, "availableGB": 0, "usagePercent": 0,
            "totalObjects": 0, "buckets": [],
            "storageType": "MinIO S3", "endpoint": cfg.get("endpoint", ""),
            "status": "disconnected",
            "error_detail": str(e),
            "snapshot_at": datetime.utcnow().isoformat(),
            "minio": {"usedGB": 0, "sizeBytes": 0, "sizeDisplay": "0.0 B",
                      "objectCount": 0, "usagePercent": 0, "status": "disconnected", "buckets": []},
            "inbox": {"usedGB": 0, "sizeBytes": 0, "sizeDisplay": "0.0 B",
                      "objectCount": 0, "usagePercent": 0, "paths": []},
            "localPath": {"usedGB": 0, "sizeBytes": 0, "sizeDisplay": "0.0 B",
                          "objectCount": 0, "usagePercent": 0, "paths": []},
        })
    finally:
        db.close()


# ──────────────────────────────────────────────
# STR-012: GET /api/storage/file/browse
# ──────────────────────────────────────────────
# ──────────────────────────────────────────────
# local_path → MinIO 정본 이관 여부 (파일 정리 지원)
# ──────────────────────────────────────────────
def _minio_import_objects(bucket, cid):
    """import/{cid}/ 아래 객체 목록 [{key,size}] — minio_object 인덱스에서 조회."""
    from sqlalchemy import text as _sql_text
    db = SessionLocal()
    try:
        rows = db.execute(_sql_text(
            "SELECT object_name, size FROM minio_object "
            "WHERE bucket = :b AND object_name LIKE :pfx"
        ), {"b": bucket, "pfx": f"import/{cid}/%"}).fetchall()
        return [{"key": r[0], "size": int(r[1] or 0)} for r in rows]
    finally:
        db.close()


def _import_match_index(objs, cid):
    """MinIO 객체 목록 → 매칭 인덱스 (by_rel, by_name) — 각각 str→set(size).

    객체 키: import/{cid}/{YYYYMMDD}/{rel_path}(정형·파일) 또는
             import/{cid}/{rel_path}(비정형). 날짜 세그먼트는 8자리 숫자.
    rel_path 전체 + basename 양쪽으로 색인 — 신규(rel 보존)·레거시(basename) 모두 매칭.
    """
    prefix = f"import/{cid}/"
    by_rel, by_name = {}, {}
    for o in objs:
        key = o.get("key", "")
        size = int(o.get("size") or 0)
        tail = key[len(prefix):] if key.startswith(prefix) else key
        first, sep, rest = tail.partition("/")
        cand = rest if (sep and len(first) == 8 and first.isdigit()) else tail
        for k in {tail, cand}:
            if k:
                by_rel.setdefault(k, set()).add(size)
        base = os.path.basename(tail)
        if base:
            by_name.setdefault(base, set()).add(size)
    return by_rel, by_name


def _import_match_status(rel, size, by_rel, by_name):
    """단일 파일의 이관 상태 — migrated / changed / not_migrated.

    migrated     : MinIO 에 같은 경로(또는 파일명)+크기 객체 존재 → 정리 안전.
    changed      : 경로는 같지만 크기가 다름 → import 후 소스 수정됨, 정리 주의.
    not_migrated : MinIO 에 사본 없음 → 아직 미이관.
    """
    rel = (rel or "").lstrip("/")
    if not rel:
        return "not_migrated"
    size = int(size or 0)
    if rel in by_rel:
        return "migrated" if size in by_rel[rel] else "changed"
    base = os.path.basename(rel)
    if base in by_name and size in by_name[base]:
        return "migrated"
    return "not_migrated"


def _table_browse(bucket, file_type, search, page, size):
    """minio_object 인덱스에서 /browse 결과 — 필터·정렬·페이지네이션을 SQL 로 처리.
    반환: (items, total). search 는 호출부에서 이미 lower-case.
    """
    from sqlalchemy import text as _sql_text
    db = SessionLocal()
    try:
        where = "bucket = :b"
        params = {"b": bucket}
        if file_type:
            where += " AND ftype = :t"
            params["t"] = file_type
        if search:
            where += " AND LOWER(name) LIKE :q"
            params["q"] = f"%{search}%"
        total = int(db.execute(_sql_text(
            f"SELECT COUNT(*) FROM minio_object WHERE {where}"), params).scalar() or 0)
        pp = dict(params, lim=size, off=max(page - 1, 0) * size)
        rows = db.execute(_sql_text(
            f"SELECT object_name, name, parent_path, ftype, extension, size, last_modified "
            f"FROM minio_object WHERE {where} "
            f"ORDER BY last_modified DESC NULLS LAST, name ASC LIMIT :lim OFFSET :off"
        ), pp).fetchall()
        items = [{
            "name": r[1], "objectName": r[0], "type": r[3] or "other",
            "extension": r[4] or "", "size": int(r[5] or 0),
            "sizeDisplay": _fmt_bytes(int(r[5] or 0)),
            "path": "/" + r[2] + "/" if r[2] else "/",
            "bucket": bucket,
            "modifiedAt": r[6].isoformat() if r[6] else None,
        } for r in rows]
        return items, total
    finally:
        db.close()


@file_bp.route("/browse", methods=["GET"])
def browse_files():
    try:
        bucket = request.args.get("bucket", MINIO_BUCKETS[0])
        file_type = request.args.get("type", "")
        search = request.args.get("search", "").lower()
        page = request.args.get("page", 1, type=int)
        size = request.args.get("size", 50, type=int)
        # minio_object 인덱스에서 SQL 로 필터·정렬·페이지네이션.
        items, total = _table_browse(bucket, file_type, search, page, size)
        return _ok(items, {"page": page, "size": size, "total": total, "bucket": bucket})
    except Exception as e:
        return _err(f"파일 목록 조회 실패: {e}", "SERVER_ERROR", 500)


@file_bp.route("/minio-events", methods=["POST"])
def minio_events():
    """MinIO Bucket Notification(webhook) 수신 → minio_object 인덱스 갱신.

    MinIO 가 호출하므로 세션이 없다 — app.py 의 require_login/enforce_rbac
    화이트리스트에 등록돼 있고, 공유 비밀(MINIO_WEBHOOK_TOKEN) Bearer 토큰으로
    검증한다. 설계: claudedocs/minio-event-index-design.md
    """
    from backend.config import MINIO_WEBHOOK_TOKEN
    if (not MINIO_WEBHOOK_TOKEN
            or request.headers.get("Authorization", "") != f"Bearer {MINIO_WEBHOOK_TOKEN}"):
        return _err("인증 실패", "UNAUTHORIZED", 401)
    try:
        from backend.services.minio_indexer import handle_events
        return _ok(handle_events(request.get_json(silent=True) or {}))
    except Exception as e:
        logger.warning("minio-events 처리 오류: %s", e)
        return _err(str(e), "SERVER_ERROR", 500)


@file_bp.route("/minio-index-status", methods=["GET"])
def minio_index_status():
    """MinIO 객체 인덱스(minio_object) 상태 — 모니터링용.

    버킷별 행수·용량·마지막 인덱싱 시각 + 마지막 reconcile 결과. 웹훅이 조용히
    끊겨도 reconcile 결과의 큰 드리프트(added/removed)로 감지할 수 있게 한다.
    """
    db = SessionLocal()
    try:
        from sqlalchemy import text as _sql_text
        rows = db.execute(_sql_text(
            "SELECT bucket, COUNT(*), COALESCE(SUM(size),0), MAX(indexed_at) "
            "FROM minio_object GROUP BY bucket ORDER BY bucket"
        )).fetchall()
        buckets = [{
            "bucket": r[0], "objectCount": int(r[1]), "sizeBytes": int(r[2] or 0),
            "sizeDisplay": _fmt_bytes(int(r[2] or 0)),
            "lastIndexedAt": r[3].isoformat() if r[3] else None,
        } for r in rows]
        last_indexed = max((b["lastIndexedAt"] for b in buckets if b["lastIndexedAt"]),
                           default=None)
        reconcile = None
        try:
            from backend.services.minio_indexer import RECONCILE_STATE_FILE
            with open(RECONCILE_STATE_FILE, "r", encoding="utf-8") as f:
                rc = _json.load(f)
            reconcile = {
                "at": datetime.utcfromtimestamp(rc["ts"]).isoformat() if rc.get("ts") else None,
                "elapsedSec": rc.get("elapsed"),
                "trigger": rc.get("trigger"),
                "result": rc.get("result"),
            }
        except Exception:
            pass
        return _ok({
            "totalObjects": sum(b["objectCount"] for b in buckets),
            "buckets": buckets,
            "lastIndexedAt": last_indexed,
            "lastReconcile": reconcile,
        })
    finally:
        db.close()


@file_bp.route("/minio-reconcile", methods=["POST"])
def minio_reconcile_trigger():
    """MinIO 인덱스 reconcile 수동 트리거 (admin) — 백그라운드 실행, 즉시 반환.

    결과는 /minio-index-status 의 lastReconcile 로 확인.
    """
    def _run():
        try:
            from backend.services.minio_indexer import run_reconcile
            run_reconcile("manual")
        except Exception as e:
            logger.warning("manual reconcile 실패: %s", e)
    threading.Thread(target=_run, name="manual-reconcile", daemon=True).start()
    return _ok({"started": True,
                "message": "reconcile 백그라운드 실행 시작 — 결과는 minio-index-status 참조"})


# ──────────────────────────────────────────────
# STR-013: GET /api/storage/file/stats
# ──────────────────────────────────────────────
@file_bp.route("/stats", methods=["GET"])
def get_file_stats():
    """파일 유형별 사용량 — minio_object 인덱스에서 유형별 GROUP BY 집계 (정본 MinIO만)."""
    try:
        # minio_object 인덱스에서 파일 유형별 집계 (GROUP BY).
        from sqlalchemy import text as _sql_text
        db = SessionLocal()
        try:
            rows = db.execute(_sql_text(
                "SELECT COALESCE(NULLIF(ftype, ''), 'other'), COUNT(*), COALESCE(SUM(size),0) "
                "FROM minio_object GROUP BY 1"
            )).fetchall()
        finally:
            db.close()
        stats = []
        for ftype, cnt, tot in rows:
            tot = int(tot or 0)
            stats.append({
                "type": ftype,
                "label": FILE_TYPE_LABELS.get(ftype, ftype),
                "count": int(cnt or 0),
                "totalSizeBytes": tot,
                "totalSizeGB": round(tot / (1024 ** 3), 4),
                "totalSizeDisplay": _fmt_bytes(tot),
            })
        stats.sort(key=lambda x: x["totalSizeBytes"], reverse=True)
        return _ok({"stats": stats, "snapshot_at": datetime.utcnow().isoformat()})
    except Exception as e:
        return _err(f"통계 조회 실패: {e}", "SERVER_ERROR", 500)


# ──────────────────────────────────────────────
# STR-014: GET/PUT /api/storage/file/cleanup-policy
# ──────────────────────────────────────────────
@file_bp.route("/cleanup-policy", methods=["GET"])
def get_cleanup_policy():
    db = _db()
    try:
        policy = db.query(FileCleanupPolicy).first()
        if not policy:
            return _ok({
                "retention_days": 90,
                "threshold_percent": 80.0,
                "enabled": False,
                "target_buckets": ["sdl-files", "sdl-archive"],
                "target_extensions": [".log", ".tmp", ".csv"],
            })
        return _ok(policy.to_dict())
    finally:
        db.close()


@file_bp.route("/cleanup-policy", methods=["PUT"])
@audit_route("storage", "storage.file.cleanup_policy.update", target_type="file_cleanup_policy",
             detail_keys=["bucket", "retentionDays", "enabled"])
def update_cleanup_policy():
    db = _db()
    try:
        body = request.get_json(force=True)

        ret = body.get("retentionDays") or body.get("retention_days")
        thr = body.get("thresholdPercent") or body.get("threshold_percent")

        if ret is not None and (int(ret) < 1 or int(ret) > 3650):
            return _err("보관 기간은 1~3650일 사이여야 합니다.", "VALIDATION")
        if thr is not None and (float(thr) < 0 or float(thr) > 100):
            return _err("임계치는 0~100% 사이여야 합니다.", "VALIDATION")

        policy = db.query(FileCleanupPolicy).first()
        if not policy:
            policy = FileCleanupPolicy()
            db.add(policy)

        if ret is not None:
            policy.retention_days = int(ret)
        if thr is not None:
            policy.threshold_percent = float(thr)
        if "enabled" in body:
            policy.enabled = bool(body["enabled"])
        if "target_buckets" in body:
            policy.target_buckets = body["target_buckets"]
        if "target_extensions" in body:
            policy.target_extensions = body["target_extensions"]

        policy.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(policy)
        return _ok(policy.to_dict())
    except Exception as e:
        db.rollback()
        return _err(str(e), "SERVER_ERROR", 500)
    finally:
        db.close()


# ──────────────────────────────────────────────
# POST /api/storage/file/upload
# ──────────────────────────────────────────────
@file_bp.route("/upload", methods=["POST"])
@audit_route("storage", "storage.file.upload", target_type="minio_object")
def upload_file():
    try:
        if "file" not in request.files:
            return _err("파일이 선택되지 않았습니다.", "VALIDATION")

        files = request.files.getlist("file")
        if not files or files[0].filename == "":
            return _err("파일이 선택되지 않았습니다.", "VALIDATION")

        bucket = request.form.get("bucket", MINIO_BUCKETS[0])
        path = request.form.get("path", "").strip("/")

        client = _get_minio()

        if not client.bucket_exists(bucket):
            return _err(f"버킷 '{bucket}'이 존재하지 않습니다.", "NOT_FOUND", 404)

        uploaded = []
        for f in files:
            filename = f.filename
            object_name = f"{path}/{filename}" if path else filename

            file_data = f.read()
            file_size = len(file_data)

            if file_size > 500 * 1024 * 1024:
                return _err(f"파일 '{filename}'의 크기가 500MB를 초과합니다.", "VALIDATION")

            client.put_object(
                bucket, object_name,
                io.BytesIO(file_data), length=file_size,
                content_type=f.content_type or "application/octet-stream",
            )
            uploaded.append({
                "name": filename,
                "objectName": object_name,
                "bucket": bucket,
                "size": file_size,
                "sizeDisplay": _fmt_bytes(file_size),
            })

        return _ok({"uploaded": uploaded, "count": len(uploaded)}), 201
    except S3Error as e:
        return _err(f"MinIO 업로드 오류: {e}", "S3_ERROR", 500)
    except Exception as e:
        return _err(f"업로드 실패: {e}", "SERVER_ERROR", 500)


# ──────────────────────────────────────────────
# DELETE /api/storage/file/delete
# ──────────────────────────────────────────────
_PREVIEW_TEXT_EXTS = {
    ".log", ".csv", ".tsv", ".txt", ".json", ".jsonl", ".ndjson",
    ".xml", ".yaml", ".yml", ".md", ".sql", ".py", ".sh", ".conf", ".ini",
    ".html", ".htm", ".css", ".js", ".ts",
}
_PREVIEW_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".svg", ".webp", ".ico"}
_PREVIEW_PDF_EXTS = {".pdf"}
_PREVIEW_BINARY_EXTS = {
    ".bin", ".dat", ".exe", ".dll", ".so", ".dylib", ".o", ".obj",
    ".pyc", ".pyo", ".class", ".jar", ".war", ".ear",
    ".tar", ".gz", ".tgz", ".bz2", ".xz", ".zip", ".7z", ".rar",
}
_PREVIEW_MAX_BYTES_DEFAULT = 256 * 1024  # 256 KB

# 텍스트 미리보기 인코딩 자동 감지 — 한글 환경(Windows Excel cp949 등) 우선.
# import_parser._detect_and_decode 와 동일한 후보 순서.
_TEXT_ENCODING_CANDIDATES = ("utf-8", "cp949", "euc-kr", "utf-8-sig", "latin-1")


def _decode_text_preview(raw):
    """미리보기 바이트를 자동 인코딩 감지하여 (text, encoding) 반환.

    부분 읽기로 멀티바이트가 잘렸을 가능성을 대비해, 모든 strict 후보 실패 시
    utf-8 errors='replace' 로 graceful fallback.
    """
    if not isinstance(raw, bytes):
        return raw, "utf-8"
    if raw[:3] == b"\xef\xbb\xbf":
        try:
            return raw.decode("utf-8-sig"), "utf-8-sig"
        except UnicodeDecodeError:
            pass
    for enc in _TEXT_ENCODING_CANDIDATES:
        try:
            return raw.decode(enc), enc
        except (UnicodeDecodeError, LookupError):
            continue
    return raw.decode("utf-8", errors="replace"), "utf-8 (replace)"


def _guess_inline_mime(ext):
    return {
        ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".gif": "image/gif", ".bmp": "image/bmp", ".svg": "image/svg+xml",
        ".webp": "image/webp", ".ico": "image/x-icon",
        ".pdf": "application/pdf",
    }.get(ext.lower(), "application/octet-stream")


# Excel 미리보기 메모리 폭증 방지 — 50MB 초과는 binary 폴백
_MAX_OPENPYXL_BYTES = 50 * 1024 * 1024
_PREVIEW_EXCEL_EXTS = (".xlsx", ".xlsm")


def _render_preview_payload(read_bytes, size, ext, raw_url, object_name, bucket,
                            sheet=None, max_bytes=_PREVIEW_MAX_BYTES_DEFAULT,
                            max_rows=100, content_type=None):
    """공유 미리보기 렌더러 — MinIO/local 양쪽이 같은 응답 shape을 만들도록 추출.

    read_bytes(n): 첫 n바이트만 lazy fetch. n=None 이면 전체 (xlsx 등).
    응답 shape는 image | pdf | table | text | binary 중 하나.
    """
    ext = (ext or "").lower()
    base = {
        "objectName": object_name,
        "bucket": bucket,
        "size": size,
        "extension": ext,
    }

    if ext in _PREVIEW_IMAGE_EXTS:
        mt = content_type if (content_type and content_type != "application/octet-stream") else _guess_inline_mime(ext)
        return {"kind": "image", **base, "mimeType": mt, "rawUrl": raw_url}

    if ext in _PREVIEW_PDF_EXTS:
        return {"kind": "pdf", **base, "rawUrl": raw_url}

    if ext in _PREVIEW_EXCEL_EXTS:
        if size and size > _MAX_OPENPYXL_BYTES:
            return {"kind": "binary", **base, "rawUrl": raw_url,
                    "message": f"엑셀 파일이 너무 큽니다 ({size//(1024*1024)} MB > 50 MB). 다운로드해서 확인하세요."}
        try:
            raw = read_bytes(None)
        except Exception as e:
            return {"kind": "binary", **base, "rawUrl": raw_url,
                    "message": f"파일 읽기 실패: {e}"}
        try:
            from io import BytesIO
            import openpyxl
            wb = openpyxl.load_workbook(BytesIO(raw), read_only=True, data_only=True)
            try:
                sheets = list(wb.sheetnames)
                cur = sheet if (sheet and sheet in sheets) else (sheets[0] if sheets else "")
                ws = wb[cur] if cur else None
                headers, rows = None, []
                if ws is not None:
                    for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        if ri == 1:
                            headers = [(str(c).strip() if c is not None else f"col_{i}") for i, c in enumerate(row)]
                            continue
                        if all(c is None for c in row):
                            continue
                        rows.append([(v.isoformat() if hasattr(v, "isoformat") else v) for v in row])
                        if len(rows) >= max_rows:
                            break
            finally:
                wb.close()
            return {"kind": "table", **base, "sheets": sheets, "currentSheet": cur,
                    "headers": headers or [], "rows": rows, "previewRows": len(rows),
                    "truncated": len(rows) >= max_rows, "rawUrl": raw_url}
        except Exception as e:
            logger.warning("excel 미리보기 실패 %s: %s", object_name, e)
            return {"kind": "binary", **base, "rawUrl": raw_url, "error": str(e)}

    if ext not in _PREVIEW_BINARY_EXTS and (ext in _PREVIEW_TEXT_EXTS or size <= 4096):
        read_size = max(1, min(size, max_bytes))
        try:
            data = read_bytes(read_size)
        except Exception as e:
            return {"kind": "binary", **base, "rawUrl": raw_url,
                    "message": f"파일 읽기 실패: {e}"}
        content, detected_encoding = _decode_text_preview(data)
        return {"kind": "text", **base,
                "previewBytes": len(data), "truncated": size > read_size,
                "content": content, "detectedEncoding": detected_encoding}

    return {"kind": "binary", **base, "rawUrl": raw_url,
            "message": "이 파일 형식은 직접 미리보기를 지원하지 않습니다. 다운로드해서 확인하세요."}


@file_bp.route("/preview", methods=["GET"])
def preview_file():
    """파일 미리보기 — 텍스트 컨텐츠 일부 또는 raw 스트리밍 URL 반환."""
    try:
        bucket = request.args.get("bucket") or MINIO_BUCKETS[0]
        object_name = request.args.get("objectName") or request.args.get("object_name")
        max_bytes = request.args.get("maxBytes", _PREVIEW_MAX_BYTES_DEFAULT, type=int)
        if not object_name:
            return _err("objectName 이 필요합니다.", "VALIDATION")

        client = _get_minio()
        try:
            stat = client.stat_object(bucket, object_name)
        except S3Error:
            return _err(f"파일을 찾을 수 없습니다: {object_name}", "NOT_FOUND", 404)

        ext = os.path.splitext(object_name)[1].lower()
        size = int(stat.size or 0)
        raw_url = (
            f"/api/storage/file/raw?bucket={bucket}"
            f"&objectName={object_name}"
        )

        def read_bytes(n):
            resp = client.get_object(bucket, object_name)
            try:
                return resp.read() if n is None else resp.read(n)
            finally:
                try:
                    resp.close()
                    resp.release_conn()
                except Exception:
                    pass

        return _ok(_render_preview_payload(
            read_bytes, size, ext, raw_url, object_name, bucket,
            sheet=request.args.get("sheet") or "",
            max_bytes=max_bytes,
            max_rows=request.args.get("maxRows", 100, type=int),
            content_type=stat.content_type,
        ))
    except Exception as e:
        return _err(str(e), "SERVER_ERROR", 500)


@file_bp.route("/raw", methods=["GET"])
def raw_file():
    """파일 원본 inline 스트리밍 (이미지/PDF 미리보기 src 용)."""
    from flask import Response, stream_with_context
    try:
        bucket = request.args.get("bucket") or MINIO_BUCKETS[0]
        object_name = request.args.get("objectName") or request.args.get("object_name")
        if not object_name:
            return _err("objectName 이 필요합니다.", "VALIDATION")
        client = _get_minio()
        try:
            stat = client.stat_object(bucket, object_name)
        except S3Error:
            return _err("파일을 찾을 수 없습니다.", "NOT_FOUND", 404)

        ext = os.path.splitext(object_name)[1].lower()
        ct = stat.content_type
        if not ct or ct == "application/octet-stream":
            ct = _guess_inline_mime(ext)

        resp = client.get_object(bucket, object_name)

        def _generate():
            try:
                while True:
                    chunk = resp.read(64 * 1024)
                    if not chunk:
                        break
                    yield chunk
            finally:
                try:
                    resp.close()
                    resp.release_conn()
                except Exception:
                    pass

        headers = {
            "Content-Disposition": f"inline; filename=\"{os.path.basename(object_name)}\"",
            "X-Content-Type-Options": "nosniff",
        }
        if stat.size:
            headers["Content-Length"] = str(stat.size)

        return Response(stream_with_context(_generate()), mimetype=ct, headers=headers)
    except Exception as e:
        return _err(str(e), "SERVER_ERROR", 500)


@file_bp.route("/delete-batch", methods=["DELETE"])
@audit_route("storage", "storage.file.delete_batch", target_type="minio_object")
def delete_files_batch():
    """여러 파일을 한 번에 삭제 (멀티 선택용).

    body: {files: [{bucket, objectName}, ...]}
    """
    try:
        body = request.get_json(force=True) or {}
        files = body.get("files") or []
        if not isinstance(files, list) or not files:
            return _err("삭제할 파일 목록이 비어있습니다.", "VALIDATION")

        client = _get_minio()
        deleted = 0
        errors = []
        for f in files:
            bucket = f.get("bucket") or MINIO_BUCKETS[0]
            object_name = f.get("objectName") or f.get("object_name")
            if not object_name:
                errors.append({"bucket": bucket, "objectName": "", "error": "objectName 누락"})
                continue
            try:
                client.remove_object(bucket, object_name)
                deleted += 1
            except S3Error as se:
                errors.append({"bucket": bucket, "objectName": object_name, "error": str(se)})
            except Exception as e:
                errors.append({"bucket": bucket, "objectName": object_name, "error": str(e)})

        return _ok({
            "requested": len(files),
            "deleted": deleted,
            "failed": len(errors),
            "errors": errors[:20],
        })
    except Exception as e:
        return _err(str(e), "SERVER_ERROR", 500)


@file_bp.route("/delete", methods=["DELETE"])
@audit_route("storage", "storage.file.delete", target_type="minio_object")
def delete_file():
    try:
        body = request.get_json(force=True)
        bucket = body.get("bucket", MINIO_BUCKETS[0])
        object_name = body.get("objectName") or body.get("object_name")

        if not object_name:
            return _err("삭제할 파일의 objectName이 필요합니다.", "VALIDATION")

        client = _get_minio()

        try:
            client.stat_object(bucket, object_name)
        except S3Error:
            return _err(f"파일을 찾을 수 없습니다: {object_name}", "NOT_FOUND", 404)

        client.remove_object(bucket, object_name)
        return _ok({"deleted": object_name, "bucket": bucket})
    except S3Error as e:
        return _err(f"MinIO 삭제 오류: {e}", "S3_ERROR", 500)
    except Exception as e:
        return _err(f"삭제 실패: {e}", "SERVER_ERROR", 500)


# ──────────────────────────────────────────────
# GET /api/storage/file/download
# ──────────────────────────────────────────────
@file_bp.route("/download", methods=["GET"])
def download_file():
    try:
        bucket = request.args.get("bucket", MINIO_BUCKETS[0])
        object_name = request.args.get("objectName") or request.args.get("object_name")

        if not object_name:
            return _err("다운로드할 파일의 objectName이 필요합니다.", "VALIDATION")

        client = _get_minio()

        try:
            stat = client.stat_object(bucket, object_name)
        except S3Error:
            return _err(f"파일을 찾을 수 없습니다: {object_name}", "NOT_FOUND", 404)

        response = client.get_object(bucket, object_name)
        file_data = response.read()
        response.close()
        response.release_conn()

        filename = os.path.basename(object_name)
        return send_file(
            io.BytesIO(file_data),
            mimetype=stat.content_type or "application/octet-stream",
            as_attachment=True,
            download_name=filename,
        )
    except S3Error as e:
        return _err(f"MinIO 다운로드 오류: {e}", "S3_ERROR", 500)
    except Exception as e:
        return _err(f"다운로드 실패: {e}", "SERVER_ERROR", 500)


# ══════════════════════════════════════════════════════════════════════
# Local Path file management — import_collector(source_mode='local_path')
# 가 가리키는 호스트 파일시스템 경로의 파일을 직접 조회/다운로드/삭제
# ══════════════════════════════════════════════════════════════════════

def _resolve_collector(db, collector_id):
    """ImportCollector 조회 + local_path 모드 검증. 실패 시 (None, err_resp)."""
    from backend.models.collector import ImportCollector
    ic = db.query(ImportCollector).get(int(collector_id))
    if not ic:
        return None, _err("Import Collector를 찾을 수 없습니다.", "NOT_FOUND", 404)
    if ic.source_mode != "local_path" or not ic.local_path:
        return None, _err("local_path 모드의 Import Collector가 아닙니다.", "INVALID_MODE", 400)
    if not os.path.isdir(ic.local_path):
        return None, _err(f"local_path가 존재하지 않습니다: {ic.local_path}", "PATH_NOT_FOUND", 404)
    return ic, None


def _safe_join(base, rel):
    """base 디렉토리 안에 머무는 경로만 허용 (path traversal 방지). 위반 시 None."""
    base_real = os.path.realpath(base)
    target = os.path.realpath(os.path.join(base, rel or ""))
    if target == base_real or target.startswith(base_real + os.sep):
        return target
    return None


@file_bp.route("/local/collectors", methods=["GET"])
def list_local_collectors():
    """local_path 모드 Import Collector 목록 (드롭다운용).

    이전엔 호출마다 모든 collector 의 os.walk(p) 로 size/count 를 계산했지만,
    NFS 위에서 162k+138k 파일 stat 이 수십 초~분 → worker timeout 위험.
    이제 file_index 테이블에서 SQL 집계 (file_indexer 백그라운드가 갱신).
    """
    from sqlalchemy import text as _sql_text
    db = SessionLocal()
    try:
        from backend.models.collector import ImportCollector
        rows = (
            db.query(ImportCollector)
              .filter(ImportCollector.source_mode == "local_path")
              .order_by(ImportCollector.id)
              .all()
        )
        # 한 번의 GROUP BY 로 모든 collector size/count 조회
        sums = {}
        if rows:
            ids = tuple(ic.id for ic in rows)
            res = db.execute(_sql_text(
                "SELECT collector_id, COUNT(*), COALESCE(SUM(size), 0) "
                "FROM file_index "
                "WHERE collector_id IN :ids AND is_dir = false "
                "GROUP BY collector_id"
            ), {"ids": ids}).fetchall()
            sums = {r[0]: (int(r[1] or 0), int(r[2] or 0)) for r in res}

        items = []
        for ic in rows:
            p = ic.local_path or ""
            # NFS isdir 도 비용 — 단순히 path 존재 여부만 확인.
            # file_index 에 행이 있으면 사실상 존재 (인덱서가 본 적 있다는 뜻).
            count, size_bytes = sums.get(ic.id, (0, 0))
            exists = count > 0 or (bool(p) and os.path.isdir(p))
            items.append({
                "id": ic.id,
                "name": ic.name,
                "path": p,
                "exists": exists,
                "sizeBytes": size_bytes,
                "sizeDisplay": _fmt_bytes(size_bytes),
                "fileCount": count,
            })
        return _ok({"items": items})
    finally:
        db.close()


@file_bp.route("/local/minio-status", methods=["POST"])
def local_minio_status():
    """local_path 파일이 MinIO 정본(import/{cid}/)에 이관됐는지 판정 — 파일 정리용.

    /local/browse 핫패스는 SQL 전용으로 유지하고, UI 가 현재 페이지의 파일
    목록을 이 엔드포인트로 보내 이관 여부 배지를 비동기로 덧입힌다.

    MinIO list 는 NFS 위에서 느리므로 import/{cid}/ 목록을 TTL 캐시(30분)하고,
    캐시 miss 시 백그라운드 워밍 + 'warming' 응답 (UI 가 잠시 후 재요청).

    Body(JSON): {collectorId, files: [{path, size}, ...]}
    Returns: {applicable, status, items: {rel_path: migrated|changed|not_migrated}, ...}
    """
    db = SessionLocal()
    try:
        body = request.get_json(silent=True) or {}
        cid = body.get("collectorId")
        files = body.get("files") or []
        if not cid:
            return _err("collectorId가 필요합니다.", "VALIDATION")
        ic, err = _resolve_collector(db, cid)
        if err:
            return err

        # target=file 만 MinIO 사본을 가짐. tsdb/rdbms 는 행으로 변환 → 해당 없음.
        target = (ic.target_type or "file").lower()
        if target != "file":
            return _ok({"applicable": False, "targetType": target,
                        "status": "ready", "items": {}})

        bucket = ic.target_bucket or "sdl-files"
        objs = _minio_import_objects(bucket, int(cid))
        by_rel, by_name = _import_match_index(objs, int(cid))
        items = {}
        migrated = changed = 0
        for f in files:
            rel = f.get("path") or ""
            st = _import_match_status(rel, f.get("size"), by_rel, by_name)
            items[rel] = st
            if st == "migrated":
                migrated += 1
            elif st == "changed":
                changed += 1
        return _ok({
            "applicable": True, "status": "ready",
            "objectCount": len(objs), "total": len(files),
            "migratedCount": migrated, "changedCount": changed,
            "items": items,
        })
    finally:
        db.close()


def _cleanup_progress_path(cid):
    return f"/tmp/sdl_cleanup_{int(cid)}.json"


def _cleanup_progress_read(cid):
    """정리 작업 진행 상태 읽기. 없거나 손상 시 None."""
    try:
        with open(_cleanup_progress_path(cid), "r", encoding="utf-8") as f:
            return _json.load(f)
    except (FileNotFoundError, ValueError):
        return None
    except Exception:
        return None


def _cleanup_progress_write(cid, data):
    """진행 상태를 작은 파일로 atomic 기록 — 컨테이너 내 모든 워커가 폴링으로 공유."""
    data["updatedAt"] = datetime.utcnow().isoformat()
    try:
        p = _cleanup_progress_path(cid)
        tmp = p + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            _json.dump(data, f, default=str)
        os.replace(tmp, p)
    except Exception as e:
        logger.warning("cleanup progress write cid=%s: %s", cid, e)


def _run_cleanup_job(cid, bucket, action, local_path, archive_subdir, name):
    """백그라운드 정리 작업 — migrated 파일 archive/delete + file_index 행 삭제.

    HTTP 요청에서 동기로 돌리면 대형 collector(수만 파일)는 gunicorn timeout(120s)
    에 워커가 죽으므로 daemon 스레드에서 처리하고, 진행 상태는 /tmp 파일로 노출한다
    (GET /local/cleanup-status 로 폴링). file_index 행은 500개 배치로 커밋해
    중간에 끊겨도 재실행이 이어서 처리한다.
    """
    from sqlalchemy import text as _sql_text
    cid = int(cid)
    db = SessionLocal()
    prog = {
        "collectorId": cid, "status": "running", "action": action,
        "total": 0, "done": 0, "cleaned": 0, "failed": 0,
        "skippedChanged": 0, "skippedNotMigrated": 0,
        "startedAt": datetime.utcnow().isoformat(), "errors": [],
    }
    try:
        # 1) 매칭 — file_index 전체를 MinIO import/{cid}/ 와 대조
        objs = _minio_import_objects(bucket, cid)
        by_rel, by_name = _import_match_index(objs, cid)
        rows = db.execute(_sql_text(
            "SELECT rel_path, size FROM file_index "
            "WHERE collector_id = :cid AND is_dir = false"
        ), {"cid": cid}).fetchall()
        migrated = []
        changed = not_migrated = 0
        for r in rows:
            st = _import_match_status(r[0], r[1], by_rel, by_name)
            if st == "migrated":
                migrated.append(r[0])
            elif st == "changed":
                changed += 1
            else:
                not_migrated += 1
        prog.update(total=len(migrated), skippedChanged=changed,
                    skippedNotMigrated=not_migrated)
        _cleanup_progress_write(cid, prog)

        # 2) archive/delete + file_index 행 삭제 (500개 배치 커밋)
        archive_base = os.path.join(local_path, archive_subdir)
        errors = []
        cleaned_total = 0
        batch = []

        def _flush_batch():
            if not batch:
                return
            part = tuple(batch)
            db.execute(_sql_text(
                "DELETE FROM file_index "
                "WHERE collector_id = :cid AND rel_path IN :paths"
            ), {"cid": cid, "paths": part})
            db.commit()
            batch.clear()

        for idx, rel in enumerate(migrated, 1):
            rel = (rel or "").strip().lstrip("/")
            full = _safe_join(local_path, rel)
            ok = False
            if not full:
                errors.append({"path": rel, "error": "경로가 base 디렉토리를 벗어납니다"})
            elif not os.path.isfile(full):
                ok = True  # 이미 사라짐 — file_index 행만 정리
            else:
                try:
                    if action == "delete":
                        os.remove(full)
                        ok = True
                    else:
                        dest = _safe_join(archive_base, rel)
                        if not dest:
                            errors.append({"path": rel, "error": "보관 경로 계산 실패"})
                        else:
                            os.makedirs(os.path.dirname(dest), exist_ok=True)
                            shutil.move(full, dest)
                            ok = True
                except OSError as e:
                    errors.append({"path": rel, "error": str(e)})
            if ok:
                batch.append(rel)
                cleaned_total += 1
            if len(batch) >= 500:
                _flush_batch()
            if idx % 200 == 0:
                prog.update(done=idx, cleaned=cleaned_total,
                            failed=len(errors), errors=errors[:20])
                _cleanup_progress_write(cid, prog)
        _flush_batch()

        prog.update(status="done", done=len(migrated), cleaned=cleaned_total,
                    failed=len(errors), errors=errors[:20],
                    finishedAt=datetime.utcnow().isoformat())
        _cleanup_progress_write(cid, prog)

        try:
            from backend.services.audit_logger import log_audit
            log_audit(
                action_type="data",
                action=f"storage.local.cleanup.{action}",
                target_type="import_collector",
                target_name=f"{cid}:{name}",
                result="success" if not errors else ("failure" if not cleaned_total else "partial"),
                detail={"collectorId": cid, "basePath": local_path, "action": action,
                        "cleaned": cleaned_total, "failed": len(errors),
                        "errors": errors[:20]},
            )
        except Exception:
            pass
        logger.info("cleanup job cid=%s done: cleaned=%d failed=%d",
                    cid, cleaned_total, len(errors))
    except Exception as e:
        logger.error("cleanup job cid=%s failed: %s", cid, e)
        prog.update(status="error", error=str(e),
                    finishedAt=datetime.utcnow().isoformat())
        _cleanup_progress_write(cid, prog)
    finally:
        db.close()


@file_bp.route("/local/cleanup-migrated", methods=["POST"])
def local_cleanup_migrated():
    """MinIO 정본에 이관 완료된 local_path 파일을 일괄 정리 (보관 이동 / 삭제).

    안전 규칙:
      - migrated(동일 경로·크기 사본이 MinIO 에 존재) 상태만 정리 대상.
      - changed / not_migrated 는 절대 건드리지 않음.
      - dryRun=true 면 대상 수만 집계 (확인 다이얼로그용, 동기 응답).
      - dryRun=false 면 백그라운드 daemon 스레드로 처리하고 즉시 반환 — 대형
        collector(수만 파일)가 gunicorn timeout(120s)에 죽지 않게.
        진행 상태는 GET /local/cleanup-status 폴링.

    Body: {collectorId, action: 'archive'|'delete', dryRun: bool}
    """
    db = SessionLocal()
    try:
        from sqlalchemy import text as _sql_text
        body = request.get_json(silent=True) or {}
        cid = body.get("collectorId")
        action = (body.get("action") or "archive").lower()
        dry_run = bool(body.get("dryRun"))
        if not cid:
            return _err("collectorId가 필요합니다.", "VALIDATION")
        if action not in ("archive", "delete"):
            return _err("action 은 archive 또는 delete 여야 합니다.", "VALIDATION")
        ic, err = _resolve_collector(db, cid)
        if err:
            return err
        if (ic.target_type or "file").lower() != "file":
            return _err("target=file 수집기만 MinIO 이관 정리를 지원합니다.", "INVALID_MODE")
        cid = int(cid)
        bucket = ic.target_bucket or "sdl-files"

        objs = _minio_import_objects(bucket, cid)

        # ── dryRun: 동기 집계 (확인 다이얼로그) ──
        if dry_run:
            by_rel, by_name = _import_match_index(objs, cid)
            rows = db.execute(_sql_text(
                "SELECT rel_path, size FROM file_index "
                "WHERE collector_id = :cid AND is_dir = false"
            ), {"cid": cid}).fetchall()
            migrated = changed = not_migrated = 0
            for r in rows:
                st = _import_match_status(r[0], r[1], by_rel, by_name)
                if st == "migrated":
                    migrated += 1
                elif st == "changed":
                    changed += 1
                else:
                    not_migrated += 1
            return _ok({
                "status": "ready", "dryRun": True, "action": action,
                "migratedCount": migrated,
                "changedCount": changed, "notMigratedCount": not_migrated,
            })

        # ── 실제 정리: 이미 실행 중이면 진행 상태 반환 ──
        prog = _cleanup_progress_read(cid)
        if prog and prog.get("status") == "running":
            # updatedAt 이 5분 넘게 멈춰 있으면 죽은 작업으로 보고 재시작 허용
            stale = True
            try:
                upd = datetime.fromisoformat(prog.get("updatedAt", ""))
                stale = (datetime.utcnow() - upd).total_seconds() > 300
            except ValueError:
                pass
            if not stale:
                return _ok({"status": "running", "progress": prog})

        # 백그라운드 작업 시작
        init = {
            "collectorId": cid, "status": "running", "action": action,
            "total": 0, "done": 0, "cleaned": 0, "failed": 0,
            "skippedChanged": 0, "skippedNotMigrated": 0,
            "startedAt": datetime.utcnow().isoformat(), "errors": [],
        }
        _cleanup_progress_write(cid, init)
        threading.Thread(
            target=_run_cleanup_job,
            args=(cid, bucket, action, ic.local_path,
                  ic.archive_subdir or ".imported", ic.name),
            name=f"cleanup-{cid}", daemon=True,
        ).start()
        return _ok({"status": "started"})
    finally:
        db.close()


@file_bp.route("/local/cleanup-status", methods=["GET"])
def local_cleanup_status():
    """이관 완료 정리 작업의 진행 상태 — UI 폴링용. Query: collectorId."""
    cid = request.args.get("collectorId", type=int)
    if not cid:
        return _err("collectorId가 필요합니다.", "VALIDATION")
    prog = _cleanup_progress_read(cid)
    if not prog:
        return _ok({"status": "idle"})
    return _ok(prog)


@file_bp.route("/local/browse", methods=["GET"])
def browse_local_files():
    """collectorId 기준 local_path 파일 트리 조회 — file_index 테이블 SQL 조회.

    이전엔 호출마다 NFS 위 os.walk 를 돌렸지만(첫 cold 5분+), 이제는 file_indexer
    백그라운드 스캐너가 PG 의 file_index 에 캐시한 결과를 ms 단위로 응답한다.
    response.indexedAt 이 stale 인지 확인하려면 /local/index-state 참고.

    Query: collectorId, path(상대), search, page, size, date_from, date_to
    """
    from sqlalchemy import text as _sql_text
    from backend.models.file_index import FileIndexState
    db = SessionLocal()
    try:
        cid = request.args.get("collectorId", type=int)
        if not cid:
            return _err("collectorId가 필요합니다.", "VALIDATION")
        ic, err = _resolve_collector(db, cid)
        if err:
            return err

        page = max(request.args.get("page", 1, type=int), 1)
        size = max(request.args.get("size", 50, type=int), 1)
        search = (request.args.get("search", "") or "").strip().lower()
        browse_path = (request.args.get("path", "") or "").strip().strip("/")
        date_from = request.args.get("date_from", "")
        date_to = request.args.get("date_to", "")

        dt_from = dt_to = None
        if date_from:
            try: dt_from = datetime.fromisoformat(date_from)
            except ValueError: pass
        if date_to:
            try: dt_to = datetime.fromisoformat(date_to)
            except ValueError: pass

        # 공통 WHERE 조건 구성 (search/페이지/cumulative 별로 재사용)
        params = {"cid": cid}
        date_cond = ""
        if dt_from is not None:
            date_cond += " AND modified_at >= :df"; params["df"] = dt_from
        if dt_to is not None:
            date_cond += " AND modified_at < :dt"; params["dt"] = dt_to

        # 검색 모드: 전체 collector 범위 + name LIKE
        # 디렉토리 탐색: parent_path = :p
        if search:
            params["q"] = f"%{search}%"
            where_files = (
                "collector_id = :cid AND is_dir = false AND "
                "(LOWER(name) LIKE :q OR LOWER(rel_path) LIKE :q)" + date_cond
            )
            where_total = where_files
        else:
            params["p"] = browse_path
            where_files = "collector_id = :cid AND parent_path = :p AND is_dir = false" + date_cond
            where_total = where_files

        # total + size 합 (현재 레벨 또는 검색결과)
        total_row = db.execute(_sql_text(
            f"SELECT COUNT(*), COALESCE(SUM(size), 0) FROM file_index WHERE {where_total}"
        ), params).first()
        total = int(total_row[0] or 0)
        total_size = int(total_row[1] or 0)

        # cumulative (collector 전체 파일)
        cum_row = db.execute(_sql_text(
            "SELECT COUNT(*), COALESCE(SUM(size), 0) FROM file_index "
            "WHERE collector_id = :cid AND is_dir = false"
        ), {"cid": cid}).first()
        cum_total = int(cum_row[0] or 0)
        cum_size = int(cum_row[1] or 0)

        # 디렉토리 목록 — 검색 모드면 비움, 일반 탐색이면 현재 레벨 directory
        dirs = []
        if not search:
            dir_rows = db.execute(_sql_text(
                "SELECT name, rel_path FROM file_index "
                "WHERE collector_id = :cid AND parent_path = :p AND is_dir = true "
                "ORDER BY name"
            ), {"cid": cid, "p": browse_path}).fetchall()
            for r in dir_rows:
                next_rel = r[1]  # 이미 base 기준 상대
                dirs.append({
                    "name": r[0],
                    "type": "directory",
                    "path": next_rel + "/",
                })

        # 파일 페이지 — 검색 결과 또는 현재 레벨
        offset = (page - 1) * size
        params_page = dict(params); params_page["lim"] = size; params_page["off"] = offset
        file_rows = db.execute(_sql_text(
            f"SELECT name, rel_path, ftype, extension, size, modified_at "
            f"FROM file_index WHERE {where_files} "
            f"ORDER BY modified_at DESC NULLS LAST, name ASC "
            f"LIMIT :lim OFFSET :off"
        ), params_page).fetchall()

        items = []
        for r in file_rows:
            mtime = r[5]
            items.append({
                "name": r[0],
                "path": r[1],
                "type": r[2] or "other",
                "extension": r[3] or "",
                "size": int(r[4] or 0),
                "sizeDisplay": _fmt_bytes(int(r[4] or 0)),
                "modifiedAt": mtime.isoformat() if mtime else None,
            })

        breadcrumb = [{"name": "root", "path": ""}]
        if browse_path:
            parts = browse_path.split("/")
            acc = ""
            for p in parts:
                acc += p + "/"
                breadcrumb.append({"name": p, "path": acc})

        state = db.query(FileIndexState).get(cid)
        index_state = state.to_dict() if state else None

        return _ok({
            "collector": {"id": ic.id, "name": ic.name, "path": ic.local_path},
            "directories": dirs,
            "items": items,
            "total": total,
            "totalSize": total_size,
            "totalSizeDisplay": _fmt_bytes(total_size),
            "cumulativeTotal": cum_total,
            "cumulativeSize": cum_size,
            "cumulativeSizeDisplay": _fmt_bytes(cum_size),
            "page": page,
            "size": size,
            "currentPath": browse_path,
            "breadcrumb": breadcrumb,
            "indexState": index_state,
        })
    finally:
        db.close()


@file_bp.route("/local/index-state", methods=["GET"])
def local_index_state():
    """파일 인덱서 상태 조회 — collectorId 단일 또는 전체."""
    db = SessionLocal()
    try:
        from backend.models.file_index import FileIndexState
        cid = request.args.get("collectorId", type=int)
        if cid:
            s = db.query(FileIndexState).get(cid)
            return _ok(s.to_dict() if s else {"collectorId": cid, "indexed": False})
        items = [s.to_dict() for s in db.query(FileIndexState).order_by(FileIndexState.collector_id).all()]
        return _ok(items)
    finally:
        db.close()


@file_bp.route("/local/reindex", methods=["POST"])
def local_reindex():
    """사용자가 수동으로 인덱싱 트리거 — 단일 collector 백그라운드 스캔.

    Query: collectorId
    """
    import threading
    cid = request.args.get("collectorId", type=int)
    if not cid:
        return _err("collectorId가 필요합니다.", "VALIDATION")

    db = SessionLocal()
    try:
        ic, err = _resolve_collector(db, cid)
        if err:
            return err
    finally:
        db.close()

    from backend.services import file_indexer
    threading.Thread(
        target=file_indexer.scan_collector,
        args=(cid,),
        name=f"manual-reindex-{cid}",
        daemon=True,
    ).start()
    return _ok({"collectorId": cid, "started": True})


@file_bp.route("/local/download", methods=["GET"])
def download_local_file():
    """local_path 파일 다운로드. Query: collectorId, path(상대)."""
    db = SessionLocal()
    try:
        cid = request.args.get("collectorId", type=int)
        rel = (request.args.get("path", "") or "").strip().lstrip("/")
        if not cid or not rel:
            return _err("collectorId와 path가 필요합니다.", "VALIDATION")
        ic, err = _resolve_collector(db, cid)
        if err:
            return err
        full = _safe_join(ic.local_path, rel)
        if not full or not os.path.isfile(full):
            return _err("파일을 찾을 수 없습니다.", "NOT_FOUND", 404)
        return send_file(full, as_attachment=True, download_name=os.path.basename(full))
    finally:
        db.close()


@file_bp.route("/local/preview", methods=["GET"])
def preview_local_file():
    """local_path 파일 미리보기 — MinIO 미리보기와 동일 응답 shape.

    Query: collectorId, path(상대), maxBytes, sheet, maxRows
    """
    db = SessionLocal()
    try:
        cid = request.args.get("collectorId", type=int)
        rel = (request.args.get("path", "") or "").strip().lstrip("/")
        if not cid or not rel:
            return _err("collectorId와 path가 필요합니다.", "VALIDATION")
        ic, err = _resolve_collector(db, cid)
        if err:
            return err
        full = _safe_join(ic.local_path, rel)
        if not full:
            return _err("경로가 base 디렉토리를 벗어납니다.", "INVALID_PATH", 400)
        if not os.path.isfile(full):
            return _err(f"파일을 찾을 수 없습니다: {rel}", "NOT_FOUND", 404)

        size = os.path.getsize(full)
        ext = os.path.splitext(full)[1].lower()
        from urllib.parse import quote
        raw_url = (
            f"/api/storage/file/local/raw?collectorId={cid}"
            f"&path={quote(rel, safe='/')}"
        )

        def read_bytes(n):
            with open(full, "rb") as fp:
                return fp.read() if n is None else fp.read(n)

        return _ok(_render_preview_payload(
            read_bytes, size, ext, raw_url, rel, "(local)",
            sheet=request.args.get("sheet") or "",
            max_bytes=request.args.get("maxBytes", _PREVIEW_MAX_BYTES_DEFAULT, type=int),
            max_rows=request.args.get("maxRows", 100, type=int),
        ))
    except Exception as e:
        return _err(str(e), "SERVER_ERROR", 500)
    finally:
        db.close()


@file_bp.route("/local/raw", methods=["GET"])
def raw_local_file():
    """local_path 파일 원본 inline 스트리밍 (image/PDF 미리보기 src 용)."""
    db = SessionLocal()
    try:
        cid = request.args.get("collectorId", type=int)
        rel = (request.args.get("path", "") or "").strip().lstrip("/")
        if not cid or not rel:
            return _err("collectorId와 path가 필요합니다.", "VALIDATION")
        ic, err = _resolve_collector(db, cid)
        if err:
            return err
        full = _safe_join(ic.local_path, rel)
        if not full:
            return _err("경로가 base 디렉토리를 벗어납니다.", "INVALID_PATH", 400)
        if not os.path.isfile(full):
            return _err("파일을 찾을 수 없습니다.", "NOT_FOUND", 404)
        ext = os.path.splitext(full)[1].lower()
        mt = _guess_inline_mime(ext)
        return send_file(full, mimetype=mt, as_attachment=False,
                         download_name=os.path.basename(full))
    finally:
        db.close()


@file_bp.route("/local/delete-batch", methods=["DELETE"])
def delete_local_files_batch():
    """local_path 파일 일괄 삭제.

    body: {collectorId: int, paths: [str, ...]}
    경로는 collector의 local_path 기준 상대경로.
    """
    db = SessionLocal()
    try:
        body = request.get_json(force=True) or {}
        cid = body.get("collectorId")
        paths = body.get("paths") or []
        if not cid or not isinstance(paths, list) or not paths:
            return _err("collectorId와 paths가 필요합니다.", "VALIDATION")
        ic, err = _resolve_collector(db, cid)
        if err:
            return err

        deleted = []
        errors = []
        for rel in paths:
            rel = (rel or "").strip().lstrip("/")
            if not rel:
                errors.append({"path": rel, "error": "빈 경로"})
                continue
            full = _safe_join(ic.local_path, rel)
            if not full:
                errors.append({"path": rel, "error": "경로가 base 디렉토리를 벗어납니다"})
                continue
            if not os.path.isfile(full):
                errors.append({"path": rel, "error": "파일이 없습니다"})
                continue
            try:
                os.remove(full)
                deleted.append(rel)
            except OSError as e:
                errors.append({"path": rel, "error": str(e)})

        # ── 감사 로그 ──
        try:
            from backend.services.audit_logger import log_audit
            log_audit(
                action_type="data",
                action="storage.local.delete",
                target_type="import_collector",
                target_name=f"{ic.id}:{ic.name}",
                result="success" if not errors else ("failure" if not deleted else "partial"),
                detail={
                    "collectorId": ic.id,
                    "basePath": ic.local_path,
                    "requested": len(paths),
                    "deleted": deleted,
                    "errors": errors[:20],
                },
            )
        except Exception:
            pass

        return _ok({
            "requested": len(paths),
            "deleted": len(deleted),
            "failed": len(errors),
            "errors": errors[:20],
        })
    finally:
        db.close()
